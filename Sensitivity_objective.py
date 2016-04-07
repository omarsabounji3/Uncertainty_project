#code pour l'analyse de sensibilite des seuils min et max
from openpyxl.reader.excel import load_workbook
from scipy import *
import numpy as np
from SALib.sample import saltelli, morris_sampling
from SALib.analyze import sobol, morris
from SALib.plotting import morris_plotting
reload(sobol)
from scipy.stats import lognorm, norm, beta, uniform
from numpy.random import dirichlet
import os
from cPickle import load, dump
from pandas import read_excel, DataFrame, concat
from copy import deepcopy
import random as rnd 
import matplotlib.pyplot as mpl
import matplotlib.pyplot as plt
from matplotlib.pyplot import title, xlabel, ylabel
from matplotlib.backends.backend_pdf import PdfPages

def read_data(filename):
    '''
Extract data from formatted xls, save as structured .pkl files for further treatment

This module converts the "'water_use_all_info.xlsx'" spreadsheet 
into a number of pickle files containing the information required to 
calculate characterization factors and carry out sensitivity and 
uncertainty analyses

Specifically:

    cell_data: 
        CU_yr_Surf, CU_yr_gw: consumptive use
        Q_90_yr: Statistical low flow
        MEAN_FG_: fraction of usage dependent on groundwater
        SUM_GWR_K: Renewabable groundwater resource
        Pi: Portion of water available of type i [0, 1]
            Note: there are (2x8=) 16 available water types,
                  all of which are independent.
                  Each water type is associated with its own alpha*
                  (and hence CF)

    country_data:
        AC: Adaptation capacity [0, 1]

    user_dist:
        [cell]x[surf, ground]x[water categories i]x[users j]

    eq_data:
        6 parameters used to calculate alpha using the sigmoid:
            A, K, Q, B, M, v
        6 parameters used for simplification of alpha
            min_1, max_1 
                Deterministic values
            max_1_low, max_1_high, min_1_low, min_1_high 
                For uncertainty and sensitivity analyses

    uncertainty_data:
        Information on distributions.
            Uncertainty types and scale factors for "cell data"
            Uncertainty types for user_distributions Uj
                Agri1, Fisheries, Dom1, Dom2, Dom3
                All Uj are beta distributed (defined in uncertainty data)
                All Uj have the same width (user_distribution)
            EF uncertainty:
                types given here, but scale factors in EF_data
                Applies to: domestic, malnutrition, WR_agriculture,
                            WR_fisheries, agri_livestock, livestock_calories
            Uncertainty of alpha simplification parameters: [min, max]

    EF_data:
        Data and uncertainty for effect factors
            Malnutrition:
                Called BHCM in Boulay et al. 2011
                Used in the calculation of Effect factors for ag and fisheries
            WR_fisheries:
                Water required to produce 1 kcal [See SI]
                Used in the calculation of Effect factor for fisheries
            WR_agriculture:
                Water required to produce 1 kcal [See SI]
                Used in the calculation of Effect factor E0,ag
            domestic:
                Effect factor for domestic (directly)
            livestock_calories:
                Ag calories per meat calories
                Used in the calculation of the Effect factor for ag
            agri_livestock:
                Portion of ag used for meat production (40%)
                Used in the calculation of the Effect factor for ag

Note:
    There are 9 users:
        Agri1, Dom1, Dom2, Dom3, Ind, Cooling, Fisheries, Hydro, Recre
        The original paper also includes Transport
        In this work, Transport, Hydro, cooling and Ind and Recre are excluded
    There are 8 water categories:
        1, 2a, 2b, 2c, 2d, 3, 4, 5
        '''

    print 'loading ' + filename + ', takes a few seconds'

    wb = load_workbook(filename)


    ws = wb.get_sheet_by_name('Cell data')
    
    print( 'reading cell data')

    cell_data = {}

    for cell_number in range(1, 809):

        cell_data[cell_number] = {}

    row_number = 0

    header = {}

    for row in ws.rows:
    #for row in []:

        row = list(row)

        row_number += 1

        if [1, 2, 3].count(row_number) == 1:

            for column in range(len(row)):

                try:

                    row[column] = str(row[column].value)

                except TypeError:

                    row[column] = ''

            header[row_number] = row

        else:

            cell_number = float(row[0].value)

            country = str(row[1].value)

            cell_data[cell_number]['Country'] = str(row[1].value)

            for column in range(2, 7):

                tag = header[3][column]

                cell_data[cell_number][tag] = float(row[column].value)

            for column in range(7, len(row)):

                tag1 = header[1][column]

                tag2 = header[2][column]

                tag3 = header[3][column]

                try:

                    cell_data[cell_number][tag1]

                except KeyError: 

                    cell_data[cell_number][tag1] = {}

                try:

                    cell_data[cell_number][tag1][tag2]

                except KeyError: 

                    cell_data[cell_number][tag1][tag2] = {}

                try:

                    element = float(row[column].value)

                except ValueError:

                    element = str(row[column].value)
                    
                except TypeError:

                    element = str(row[column].value)
                
                cell_data[cell_number][tag1][tag2][tag3] = element

    

    ws = wb.get_sheet_by_name('Country data')
    print ('reading Country data')
    country_data = {}

    row_number = 0

    for row in ws.rows:
    #for row in []:

        row = list(row)

        row_number += 1

        if row_number != 1:

            country = str(row[0].value)

            try:

                AC = float(row[1].value)
                Potable= float(row[2].value)
                Sanitation=float(row[3].value)
                GNI=float(row[4].value)

            except ValueError:

                AC = 'NA'
                Potable='NA'
                Sanitation='NA'
                GNI='NA'
                
            country_data[country] = {}

            country_data[country]['AC'] = AC
            country_data[country]['Potable'] = Potable
            country_data[country]['Sanitation'] = Sanitation
            country_data[country]['GNI'] = GNI


    

    ws = wb.get_sheet_by_name('User distribution')
    print ('reading User distribution')
    user_dist = {}

    row_number = 0

    header = {}

    for row in ws.rows:
    #for row in []:
        
        row = list(row
        )

        row_number += 1

        if [1, 2, 3].count(row_number) == 1:

            for column in range(len(row)):

                try:

                    row[column] = str(row[column].value)

                except TypeError:

                    row[column] = ''

            header[row_number] = row

        else:

            cell_number = int(row[0].value)

            user_dist[cell_number] = {}

            for column in range(1,len(row)):

                ground_surface = header[1][column]

                try:

                    user_dist[cell_number][ground_surface]

                except KeyError:

                    user_dist[cell_number][ground_surface] = {}

                water_type = header[2][column]

                try:

                    user_dist[cell_number][ground_surface][water_type]

                except KeyError:

                    user_dist[cell_number][ground_surface][water_type] = {}

                user = header[3][column]

                if user != 'Hydro' and user != 'Recre':

                    try:

                        data = float(row[column].value)

                    except TypeError:

                        data = 'NA'

                    user_dist[cell_number][ground_surface][water_type][user] = data
                    
    ws = wb.get_sheet_by_name('Uncertainty data')
    print ('reading uncertainty data')
    uncertainty_data = {}
    
    for row in ws.rows:
    #for row in []:
        
        variable = row[0]
        data_type = row[1]
        data = row[2]
    
        variable = str(variable.value)
    
        data_type = str(data_type.value)
        
        if data.value == None :  
            break
        try:
            data = float(data.value)
        except ValueError:
            
            data= str(data.value)
    
        try:
        
        
            uncertainty_data[variable]
    
        except KeyError:
        
            uncertainty_data[variable] = {}
            
        uncertainty_data[variable][data_type] = data
        
    ws = wb.get_sheet_by_name('Equation data')
    print( 'reading equation data')
    eq_data = {}
        
    for row in ws.rows:
        
        variable= row[0]
        
        data=row[1]
        
        variable = str(variable.value)
        
        data = float(data.value)
        
        eq_data[variable] = data
    
    ws = wb.get_sheet_by_name('EF')
    print ('reading EF')
    EF_data = {}
    for row in ws.rows:
    #for row in []:
        
        variable = row[0]
        data_type = row[1]
        data = row[2]
        unit = row[3]
        
        variable = str(variable.value)
        
        data_type = str(data_type.value)
        
        data = float(data.value)
        try:
            
            EF_data[variable]
        except KeyError:
            
            EF_data[variable] = {}
            
        EF_data[variable][data_type] = data

    return cell_data, country_data, user_dist, eq_data, uncertainty_data, EF_data
def save_variable(all_info):
    ''' 
    This function saves the differents contents of the "water_use_all_info.xlsx" spreadsheets in .pkl files'''
    #all_info = [
    #           [path, variable_name, variable], 
    #           [path, variable_name, variable]
    #           ]
    for path, variable_name, variable in all_info:
        filename = os.path.join(path, variable_name + '.pkl')
        file = open(filename, 'wb')
        dump(variable, file)
        file.close()
def load_variables(all_info):
    '''This module is used to load data from the .pkl files'''
    #all_info = [[path, variable_name], [path, variable_name], ...]
    variables = []
    for path, variable_name in all_info:
        filename = path + '/' + variable_name + '.pkl'
        variables.append(load(open(filename, 'rb')))
    return variables
def build_file_list(dirpath, extension = None):
    '''This function is used to give a list of pickles files that are in the folder'''
    pre_list = os.listdir(dirpath)
    if extension == None:
        filelist = [filename for filename in pre_list 
                if os.path.isfile(os.path.join(dirpath, filename)) and '~' not in filename]
    else:
        filelist = [filename for filename in pre_list 
                if os.path.isfile(os.path.join(dirpath, filename))
                and filename.split('.')[-1].lower() == extension and '~' not in filename]
    
    return filelist
def calculate_CF(parameters):
    '''The main purpose of this function is to calculate the CF using either deterministic or probabilistic values.

    The input parameters have to be in a sort of configuration:

        -The User distributions called Uj have to be in a dictionnary and they are aggregated in categories: 
            -the domestic users 'Dom1', 'Dom2', 'Dom3' are classified in the 'domestic' category.
            -the agriculture user Agri1 is classified in the 'WR_agriculture' category.
            -the fisheries user Fisheries are classified in the 'WR_fisheries' category.
        -The effect factors are calculated '''
    users_aggregation = {}
    users_aggregation['WR_agriculture'] = ['Agri1']
    users_aggregation['domestic'] = ['Dom1', 'Dom2', 'Dom3']
    users_aggregation['WR_fisheries'] = ['Fisheries']
    deterministic_CF = 0.
    alpha = 'NA'
    alpha_star= 'NA'
    if parameters['AC'] != 1.:
        if ((parameters['SUM_GWR_K'] == 0. and ground_surface == 'groundwater') or 
                (parameters['Q_90_yr'] == 0. and ground_surface == 'surface')):
            deterministic_CF = 'NA'
        else:
            if ground_surface == 'groundwater':
                alpha_star = parameters['CU_yr_gw']* parameters['MEAN_FG_'] / parameters['SUM_GWR_K'] / parameters['Pi']
            else:
                alpha_star = parameters['CU_yr_Surf']*(1.-parameters['MEAN_FG_'])/ parameters['Q_90_yr'] / parameters['Pi']
#            print alpha_star
            if alpha_star < parameters['min']:
                alpha = 0.
            elif alpha_star > parameters['max']:
                alpha = 1.
            else:
                alpha = (alpha_star-parameters['min'])/(parameters['max'] - parameters['min'])
                #alpha = eq_data['K'] / ((1. + (eq_data['Q'] * exp(-eq_data['B'] * (alpha_star - eq_data['M'])))**(1. / eq_data['v'])))
            if alpha == 0. :
                deterministic_CF= 0.
            else:
                for user_general in users_aggregation:
    #                print user_general
                    if user_general == 'domestic':
                        EF = parameters[user_general]
                    elif user_general == 'WR_fisheries':
                        EF = parameters['malnutrition'] / parameters[user_general]
                    else:
                        EF_uncorrected = parameters['malnutrition'] / parameters[user_general]
                        EF = (((1. - parameters['agri_livestock']) * EF_uncorrected) + 
                                (parameters['agri_livestock']* EF_uncorrected / parameters['livestock_calories']))
                    for user_specific in users_aggregation[user_general]:
                        deterministic_CF += alpha * (1. - parameters['AC']) * EF * parameters[user_specific]
    return deterministic_CF, alpha, alpha_star
def create_rv(parameters_deterministic_values, parameters_uncertainty_values):
    RVs = {}
    all_data_available= True
    for parameter in parameters_uncertainty_values:
        if parameters_deterministic_values[parameter]==0.:
            RVs[parameter]= uniform(loc = 0., scale = 0.) 
        elif parameters_uncertainty_values[parameter]['distribution'] == 'uniform':
            if 'width' in parameters_uncertainty_values[parameter]:
                loc = parameters_deterministic_values[parameter] - parameters_uncertainty_values[parameter]['width']
                if loc< 0.:
                    1/0
                scale = 2.*parameters_uncertainty_values[parameter]['width']
            else:
                loc= parameters_uncertainty_values[parameter]['min']
                scale = parameters_uncertainty_values[parameter]['max'] -parameters_uncertainty_values[parameter]['min']
            RVs[parameter] = uniform(loc = loc, scale = scale) 
        elif parameters_uncertainty_values[parameter]['distribution'] == 'normal':
            RVs[parameter] = norm(loc = parameters_deterministic_values[parameter] , scale = parameters_uncertainty_values[parameter]['std'])
        elif parameters_uncertainty_values[parameter]['distribution'] == 'lognormal':
            RVs[parameter] = fit_lognormal(parameters_deterministic_values[parameter], parameters_uncertainty_values[parameter]['GSD2'])
        elif parameters_uncertainty_values[parameter]['distribution'] == 'beta':
#            RVs[parameter] = create_beta(parameters_deterministic_values[parameter], parameters_uncertainty_values[parameter])
            if 'width' in parameters_uncertainty_values[parameter]:
                loc = parameters_deterministic_values[parameter] - parameters_uncertainty_values[parameter]['width']
                if loc< 0.:
                    loc=0.
                scale = 2.*parameters_uncertainty_values[parameter]['width']
            else:
                loc= parameters_uncertainty_values[parameter]['min']
                scale = parameters_uncertainty_values[parameter]['max'] -parameters_uncertainty_values[parameter]['min']
            RVs[parameter] = uniform(loc = loc, scale = scale) 
    return RVs, all_data_available
def unif_dist_parameters(parameters_deterministic_values, parameters_uncertainty_values):
    loc=0.
    scale=0.
    for parameter in parameters_uncertainty_values:
        if parameters_deterministic_values==0.:
            loc = 0. 
            scale = 0.
        elif parameters_uncertainty_values['distribution'] == 'uniform':
            if 'width' in parameters_uncertainty_values:
                loc = parameters_deterministic_values - parameters_uncertainty_values['width']
                if loc< 0.:
                    1/0
                scale = parameters_deterministic_values+parameters_uncertainty_values['width']
            else:
                loc= parameters_uncertainty_values['min']
                scale = parameters_uncertainty_values['max']
        elif parameters_uncertainty_values['distribution'] == 'beta':
#            RVs[parameter] = create_beta(parameters_deterministic_values[parameter], parameters_uncertainty_values[parameter])
            if 'width' in parameters_uncertainty_values:
                loc = parameters_deterministic_values - parameters_uncertainty_values['width']
                if loc< 0.:
                    loc=0.
                scale = parameters_deterministic_values+parameters_uncertainty_values['width']
            else:
                loc= parameters_uncertainty_values['min']
                scale = parameters_uncertainty_values['max']
    return [loc,scale]
def alpha_star_analysis(parameters_uncertainty_values, min_alpha_stars, max_alpha_stars):
    contents = []
    if parameters_uncertainty_values['min']['min'] > min_alpha_stars and parameters_uncertainty_values['min']['min'] < max_alpha_stars:
        contents.append(1)
    if parameters_uncertainty_values['min']['max'] > min_alpha_stars and parameters_uncertainty_values['min']['max'] < max_alpha_stars:
        contents.append(2)
    if parameters_uncertainty_values['max']['min'] > min_alpha_stars and parameters_uncertainty_values['max']['min'] < max_alpha_stars:
        contents.append(3)
    if parameters_uncertainty_values['max']['max'] > min_alpha_stars and parameters_uncertainty_values['max']['max'] < max_alpha_stars:
        contents.append(4)
    return contents
def fit_lognormal(average, GSD2):
    s = np.log(GSD2**.5)
    rv = lognorm(s, scale = average)
    rv.ppf([.0225, .5, .9775])
    return rv
def create_beta(average, parameters_uncertainty_values):
    if 'width' in parameters_uncertainty_values:
        minimum = average - parameters_uncertainty_values['width']
        maximum = average + parameters_uncertainty_values['width']
    else:
        minimum = parameters_uncertainty_values['min']
        maximum = parameters_uncertainty_values['max']
    if minimum < 0.:
        minimum =0.
    p975 = minimum + (maximum - minimum)*.975
    assert maximum > minimum, 'minimum should be lower than maximum'
    assert average > minimum, 'average should be higher than minimum'
    assert average < maximum, 'average should be lower than maximum'
    assert p975 > average, 'p975 should be higher than average'
    mu = (average - minimum) / (maximum - minimum)
    ba_ratio = (1 - mu)/mu
    scale = maximum - minimum
    #initial guess
    a = 1.
    b = a*ba_ratio
    threshold = .01
    direction = 0
    delta = .2
    counter = 0
    while 1:
        counter += 1
        rv = beta(a, b, loc = minimum, scale = scale)
        if abs(rv.ppf(.975)/p975) < 1.+threshold and abs(rv.ppf(.975)/p975 > 1.-threshold):
            break
        elif rv.ppf(.975) > p975:
            if direction == -1:
                delta = delta/2.
            direction = 1
        else:
            if direction == 1:
                delta = delta/2.
            direction = -1
        if counter > 1000:
            print (counter)
        a += direction*delta
        b = a*ba_ratio
    return rv
def create_deltas(uncertainty_data, EF_data):
    deltas = {}
    deltas['malnutrition']= {}
    deltas['malnutrition']['min']=1./EF_data['malnutrition']['GSD2']
    deltas['malnutrition']['max']=EF_data['malnutrition']['GSD2']
    deltas['WR_fisheries']={}
    deltas['WR_fisheries']['min']=EF_data['WR_fisheries']['min']/EF_data['WR_fisheries']['average']
    deltas['WR_fisheries']['max']=EF_data['WR_fisheries']['max']/EF_data['WR_fisheries']['average']
    deltas['WR_agriculture']= {}
    deltas['WR_agriculture']['min']=EF_data['WR_agriculture']['min']/EF_data['WR_agriculture']['average']
    deltas['WR_agriculture']['max']=EF_data['WR_agriculture']['max']/EF_data['WR_agriculture']['average']
    deltas['domestic']= {}
    deltas['domestic']['min']=1./EF_data['domestic']['GSD2']
    deltas['domestic']['max']=EF_data['domestic']['GSD2']
    deltas['livestock_calories']= {}
    deltas['livestock_calories']['min']=EF_data['livestock_calories']['min']/EF_data['livestock_calories']['average']
    deltas['livestock_calories']['max']=EF_data['livestock_calories']['max']/EF_data['livestock_calories']['average']
    deltas['agri_livestock']= {}
    deltas['agri_livestock']['min']=(EF_data['agri_livestock']['average']-EF_data['agri_livestock']['width'])/EF_data['agri_livestock']['average']
    deltas['agri_livestock']['max']=(EF_data['agri_livestock']['average']+EF_data['agri_livestock']['width'])/EF_data['agri_livestock']['average']
    deltas['CU_yr_Surf']= {}
    deltas['CU_yr_Surf']['min']=1./uncertainty_data['CU_yr_Surf']['GSD2']
    deltas['CU_yr_Surf']['max']=uncertainty_data['CU_yr_Surf']['GSD2']
    deltas['CU_yr_gw']= {}
    deltas['CU_yr_gw']['min']=1./uncertainty_data['CU_yr_Surf']['GSD2']
    deltas['CU_yr_gw']['max']=uncertainty_data['CU_yr_Surf']['GSD2']
    deltas['Q_90_yr']= {}
    deltas['Q_90_yr']['min']=1./uncertainty_data['Q_90_yr']['GSD2']
    deltas['Q_90_yr']['max']=uncertainty_data['Q_90_yr']['GSD2']
    deltas['SUM_GWR_K']= {}
    deltas['SUM_GWR_K']['min']=1./uncertainty_data['SUM_GWR_K']['GSD2']
    deltas['SUM_GWR_K']['max']=uncertainty_data['SUM_GWR_K']['GSD2']
    deltas['MEAN_FG_']= {}
    deltas['MEAN_FG_']['min']=1. - uncertainty_data['MEAN_FG_']['width']
    deltas['MEAN_FG_']['max']=1. + uncertainty_data['MEAN_FG_']['width']
    deltas['Pi']= {}
    deltas['Pi']['min']=1. - uncertainty_data['Pi']['width']
    deltas['Pi']['max']=1. + uncertainty_data['Pi']['width']
    deltas['Cooling']= {}
    deltas['Cooling']['min']=1. - uncertainty_data['user_distribution']['width']
    deltas['Cooling']['max']=1. + uncertainty_data['user_distribution']['width']
    deltas['Agri1']= {}
    deltas['Agri1']['min']=1. - uncertainty_data['user_distribution']['width']
    deltas['Agri1']['max']=1. + uncertainty_data['user_distribution']['width']
    deltas['Ind']= {}
    deltas['Ind']['min']=1. - uncertainty_data['user_distribution']['width']
    deltas['Ind']['max']=1. + uncertainty_data['user_distribution']['width']
    deltas['Dom1']= {}
    deltas['Dom1']['min']=1. - uncertainty_data['user_distribution']['width']
    deltas['Dom1']['max']=1. + uncertainty_data['user_distribution']['width']
    deltas['Dom2']= {}
    deltas['Dom2']['min']=1. - uncertainty_data['user_distribution']['width']
    deltas['Dom2']['max']=1. + uncertainty_data['user_distribution']['width']
    deltas['Dom3']= {}
    deltas['Dom3']['min']=1. - uncertainty_data['user_distribution']['width']
    deltas['Dom3']['max']=1. + uncertainty_data['user_distribution']['width']
    deltas['Fisheries']= {}
    deltas['Fisheries']['min']=1. - uncertainty_data['user_distribution']['width']
    deltas['Fisheries']['max']=1. + uncertainty_data['user_distribution']['width']
    deltas['AC']= {}
    deltas['AC']['min']=0.9
    deltas['AC']['max']=1.1
    deltas['max']= {}
    deltas['max']['min']=eq_data['max_1_low']/eq_data['max_1']
    deltas['max']['max']=eq_data['max_1_high']/eq_data['max_1']
    deltas['min']= {}
    deltas['min']['min']=eq_data['min_1_low']/eq_data['min_1']
    deltas['min']['max']=eq_data['min_1_high']/eq_data['min_1']
    return deltas
def select_SA_sample(sample, all_data):
    selected_data= {}
    for p in parameters_deterministic_values:
        if p in sample:
            selected_data[p]= sample[p]
        else:
            selected_data[p]=parameters_deterministic_values[p]
    return selected_data
def create_sample(RVs, nb_iteration):
 #create a sample of nb_iteration for each parameter
    sample = {}
    for p in RVs:
        sample[p] = RVs[p].rvs(nb_iteration)
    return sample
def select_sample(sample, all_data, n):
    selected_data= {}
    for p in parameters_deterministic_values:
        if p in sample:
            selected_data[p]= sample[p][n]
        else:
            selected_data[p]=parameters_deterministic_values[p]
    return selected_data
def stats_on_CF(CFs, alphas, alpha_stars):
    m = np.mean(CFs)
    m_alpha = np.mean(alphas)
    m_alpha_star=np.mean(alpha_stars)
    m = np.mean(CFs)
    mini=np.min(CFs)
    maxi=np.max(CFs)
    confidence_95 = [np.percentile(CFs, 2.5), np.percentile(CFs, 97.5)]
    median = np.percentile(CFs, 50.)
    min_alphas=np.min(alphas)
    min_alpha_stars=np.min(alpha_stars)
    max_alphas=np.max(alphas)
    max_alpha_stars=np.max(alpha_stars)
    return m, mini, maxi, confidence_95, median, min_alpha_stars, max_alpha_stars, min_alphas, max_alphas, m_alpha,m_alpha_star 
def extract_data(cell_ID, ground_surface, water_type, cell_data, country_data, eq_data,EF_data):
    all_data_available = False
    while 1:
        parameters = {}
        if country_data[cell_data[cell_ID]['Country']]['AC'] == 'NA':
            parameters['AC'] = country_data[cell_data[cell_ID]['Country']]['AC']
            break
        parameters['AC'] = country_data[cell_data[cell_ID]['Country']]['AC']
        parameters['CU_yr_gw'] = cell_data[cell_ID]['CU_yr_gw']
        parameters['SUM_GWR_K'] = cell_data[cell_ID]['SUM_GWR_K']
        parameters['CU_yr_Surf'] = cell_data[cell_ID]['CU_yr_Surf']
        parameters['Q_90_yr'] = cell_data[cell_ID]['Q_90_yr']
        parameters['MEAN_FG_']= cell_data[cell_ID]['MEAN_FG_']
        if cell_data[cell_ID]['Pi'][ground_surface][water_type] in ['NA', '99999']:
            parameters['Pi']=cell_data[cell_ID]['Pi'][ground_surface][water_type]
            break
        parameters['Pi']=cell_data[cell_ID]['Pi'][ground_surface][water_type]
        parameters.update(user_dist[cell_ID][ground_surface][water_type])
        all_data_available = True
        break
    parameters['min'] = eq_data['min_1']
    parameters['max'] = eq_data['max_1']
    parameters['malnutrition']=EF_data['malnutrition']['average']
    parameters['domestic'] = EF_data['domestic']['average']
    parameters['WR_agriculture'] = EF_data['WR_agriculture']['average']
    parameters['WR_fisheries']= EF_data['WR_fisheries']['average']
    parameters['agri_livestock'] = EF_data['agri_livestock']['average']
    parameters['livestock_calories'] = EF_data['livestock_calories']['average']
    return parameters, all_data_available
def extract_uncertainty_info(cell_ID, ground_surface, water_type, cell_data, country_data, eq_data, EF_data, uncertainty_data):
	#put in a dictionary like this:

    parameters_uncertainty_values = {}
    parameters_uncertainty_values['malnutrition'] = {}
    parameters_uncertainty_values['domestic'] = {}
    parameters_uncertainty_values['CU_yr_Surf'] = {}
    parameters_uncertainty_values['CU_yr_gw'] = {}
    parameters_uncertainty_values['Q_90_yr'] = {}
    parameters_uncertainty_values['SUM_GWR_K'] = {}
    parameters_uncertainty_values['agri_livestock'] = {}
    parameters_uncertainty_values['MEAN_FG_'] = {}
    parameters_uncertainty_values['Pi'] = {}
    parameters_uncertainty_values['Fisheries'] = {}
    parameters_uncertainty_values['Dom1'] = {}
    parameters_uncertainty_values['Dom2'] = {}
    parameters_uncertainty_values['Dom3'] = {}
    parameters_uncertainty_values['Agri1'] = {}
    parameters_uncertainty_values['min'] = {}
    parameters_uncertainty_values['max'] = {}
    parameters_uncertainty_values['WR_fisheries'] = {}
    parameters_uncertainty_values['WR_agriculture'] = {}
    parameters_uncertainty_values['livestock_calories'] = {}
    
    
    parameters_uncertainty_values['malnutrition']['GSD2'] = EF_data['malnutrition']['GSD2']
    parameters_uncertainty_values['domestic']['GSD2']=EF_data['domestic']['GSD2']
    parameters_uncertainty_values['CU_yr_Surf']['GSD2']=uncertainty_data['CU_yr_Surf']['GSD2']
    parameters_uncertainty_values['CU_yr_gw']['GSD2']=uncertainty_data['CU_yr_gw']['GSD2']
    parameters_uncertainty_values['Q_90_yr']['GSD2']=uncertainty_data['Q_90_yr']['GSD2']
    parameters_uncertainty_values['SUM_GWR_K']['GSD2']=uncertainty_data['SUM_GWR_K']['GSD2']

    parameters_uncertainty_values['agri_livestock']['width']=EF_data['agri_livestock']['width']
    parameters_uncertainty_values['MEAN_FG_']['width']=uncertainty_data['MEAN_FG_']['width']
    parameters_uncertainty_values['Pi']['width']=uncertainty_data['Pi']['width']
    parameters_uncertainty_values['Fisheries']['width']=uncertainty_data['user_distribution']['width']
    parameters_uncertainty_values['Dom1']['width']=uncertainty_data['user_distribution']['width']
    parameters_uncertainty_values['Dom2']['width']=uncertainty_data['user_distribution']['width']
    parameters_uncertainty_values['Dom3']['width']=uncertainty_data['user_distribution']['width']
    parameters_uncertainty_values['Agri1']['width']=uncertainty_data['user_distribution']['width']    
    
    parameters_uncertainty_values['min']['min']=eq_data['min_1_low']
    parameters_uncertainty_values['min']['max']=eq_data['min_1_high']
    parameters_uncertainty_values['max']['min']=eq_data['max_1_low']
    parameters_uncertainty_values['max']['max']=eq_data['max_1_high']
    parameters_uncertainty_values['WR_fisheries']['min']=EF_data['WR_fisheries']['min']
    parameters_uncertainty_values['WR_fisheries']['max']=EF_data['WR_fisheries']['max']
    parameters_uncertainty_values['WR_agriculture']['min']=EF_data['WR_agriculture']['min']
    parameters_uncertainty_values['WR_agriculture']['max']=EF_data['WR_agriculture']['max']
    parameters_uncertainty_values['livestock_calories']['min']=EF_data['livestock_calories']['min']
    parameters_uncertainty_values['livestock_calories']['max']=EF_data['livestock_calories']['max']
    
    
    parameters_uncertainty_values['Pi']['distribution']=uncertainty_data['Pi']['distribution']
    parameters_uncertainty_values['MEAN_FG_']['distribution']=uncertainty_data['MEAN_FG_']['distribution']
    parameters_uncertainty_values['SUM_GWR_K']['distribution']=uncertainty_data['SUM_GWR_K']['distribution']
    parameters_uncertainty_values['Q_90_yr']['distribution']=uncertainty_data['Q_90_yr']['distribution']
    parameters_uncertainty_values['CU_yr_gw']['distribution']=uncertainty_data['CU_yr_gw']['distribution']
    parameters_uncertainty_values['CU_yr_Surf']['distribution']=uncertainty_data['CU_yr_Surf']['distribution']
    parameters_uncertainty_values['livestock_calories']['distribution']=uncertainty_data['livestock_calories']['distribution']
    parameters_uncertainty_values['agri_livestock']['distribution']=uncertainty_data['agri_livestock']['distribution']
    parameters_uncertainty_values['domestic']['distribution']=uncertainty_data['domestic']['distribution']
    parameters_uncertainty_values['WR_agriculture']['distribution']= uncertainty_data['WR_agriculture']['distribution']
    parameters_uncertainty_values['WR_fisheries']['distribution']=uncertainty_data['WR_fisheries']['distribution']
    parameters_uncertainty_values['malnutrition']['distribution']=uncertainty_data['malnutrition']['distribution']
    parameters_uncertainty_values['Fisheries']['distribution']=uncertainty_data['Fisheries']['distribution']
    parameters_uncertainty_values['Dom1']['distribution']=uncertainty_data['Dom1']['distribution']
    parameters_uncertainty_values['Dom2']['distribution']=uncertainty_data['Dom2']['distribution']
    parameters_uncertainty_values['Dom3']['distribution']=uncertainty_data['Dom3']['distribution']
    parameters_uncertainty_values['Agri1']['distribution']=uncertainty_data['Agri1']['distribution']    
    parameters_uncertainty_values['min']['distribution']=uncertainty_data['min']['distribution']
    parameters_uncertainty_values['max']['distribution']=uncertainty_data['max']['distribution']    
    

    return parameters_uncertainty_values
def dirichlet_sample(sample, parameters_deterministic_values):
    det_shares=[]
    CF_users_sample={}
    User={}
    User_affected={}
    User_affected_CF={}
    CF_users_positions=[]
    User['Ind']=parameters_deterministic_values['Ind']
    User['Cooling']=parameters_deterministic_values['Cooling']
    User['Dom1']=parameters_deterministic_values['Dom1']
    User['Dom2']=parameters_deterministic_values['Dom2']
    User['Dom3']=parameters_deterministic_values['Dom3']
    User['Agri1']=parameters_deterministic_values['Agri1']
    width=[]
    scaling =[1100, 100]
    for i in User:
        if User[i]!=0. :
            User_affected[i]=User[i]
    for i in User_affected:
        if i in ['Dom1','Dom2','Dom3','Agri1']:
            del sample[i]
            User_affected_CF[i]=User_affected[i]
        det_shares.append(User_affected[i]) 
    det_shares=np.array(det_shares)
    s = [dirichlet(det_shares*scaling[0], nb_iteration), 
         dirichlet(det_shares*scaling[1], nb_iteration)]
    det_shares_list=list(det_shares)
    for i in User_affected_CF:
        CF_users_positions.append(det_shares_list.index(User_affected_CF[i]))
        CF_users_sample[i]=[]
    for i in range(len(s[0])):
        for j in CF_users_positions:
            CF_users_sample[User_affected.keys()[j]].append(s[0][i][j])                
    sample.update(CF_users_sample)
    for i in CF_users_sample:
        width.append(np.percentile(CF_users_sample[i], 97.5)-np.percentile(CF_users_sample[i], 0.025))
    return sample
    
folder = '/Users/Omar/Desktop/Projet_recherche/Water Use/Sensitivity analysis/Code'
#folder = r'C:\python\water_use_MC_2\src'
force_read_excel=False
if 'cell_data.pkl' in build_file_list(folder) and not force_read_excel :
    var_names = ['cell_data', 'country_data', 'user_dist','eq_data', 'uncertainty_data', 'EF_data']
    print ('loading pickle')
    all_info = zip([folder]*len(var_names), var_names)
    cell_data, country_data, user_dist,eq_data, uncertainty_data, EF_data = load_variables(all_info)
    print ('finish loading pickle' )
else:
    filename = 'water_use_all_info.xlsx'
    cell_data, country_data, user_dist,eq_data, uncertainty_data, EF_data = read_data(filename)
    all_info = [[folder, 'cell_data', cell_data], 
                [folder, 'country_data', country_data], 
                [folder, 'user_dist', user_dist], 
                [folder, 'eq_data', eq_data], 
                [folder, 'uncertainty_data', uncertainty_data], 
                [folder, 'EF_data', EF_data]]
    save_variable(all_info)
Sobol_result = DataFrame()
Morris_result= DataFrame()

nb_iteration = 1000
Cells={}
#for cell_ID in [7]:
#    Cells[cell_ID]={}
#    for ground_surface in ['surface']:
#        Cells[cell_ID][ground_surface]={}
#        for water_type in ['2c']:
#            Cells[cell_ID][ground_surface][water_type]={}
for cell_ID in cell_data:
    Cells[cell_ID]={}
    for ground_surface in cell_data[cell_ID]['Pi']:
        Cells[cell_ID][ground_surface]={}
        for water_type in cell_data[cell_ID]['Pi'][ground_surface]:
            Cells[cell_ID][ground_surface][water_type]={}
        #calculate if all data available
            print ('Cell: %s, water: %s %s' % (cell_ID, water_type, ground_surface))
            parameters_deterministic_values, all_data_available = extract_data(cell_ID, ground_surface, water_type, cell_data, country_data, eq_data,EF_data)
            Cells[cell_ID][ground_surface][water_type]['data availability']=all_data_available
            Cells[cell_ID][ground_surface][water_type]['parameters deterministic values']=parameters_deterministic_values
            to_add = {0: {'cell_ID': cell_ID, 
								'ground_surface': ground_surface, 
								'water_type': water_type, 
								'country': cell_data[cell_ID]['Country'], 
								}}
            if all_data_available:
                deterministic_CF, deterministic_alpha, deterministic_alpha_star = calculate_CF(parameters_deterministic_values)
                Cells[cell_ID][ground_surface][water_type]['deterministic CF']=deterministic_CF
                parameters_uncertainty_values = extract_uncertainty_info(cell_ID, ground_surface, water_type, cell_data, country_data, eq_data, EF_data, uncertainty_data)
                Cells[cell_ID][ground_surface][water_type]['probabilistic parameters']=parameters_uncertainty_values.keys()
                Cells[cell_ID][ground_surface][water_type]['parameters distribution']=parameters_uncertainty_values
                RVs, all_data_available = create_rv(parameters_deterministic_values, parameters_uncertainty_values)
                sample = create_sample(RVs, nb_iteration)
                deltas= create_deltas(uncertainty_data, EF_data)
#                sample=dirichlet_sample(sample, parameters_deterministic_values)
                CFs = []
                alpha_stars = []
                alphas = []
                CF_sobol_ratio=[]
                CF_Morris_ratio=[]
                for iteration in range(nb_iteration):        
                    selected_sample = select_sample(sample, parameters_deterministic_values, iteration)
                    CF, alpha, alpha_star = calculate_CF(selected_sample)
                    CFs.append(CF)
                    alpha_stars.append(alpha_star)
                    alphas.append(alpha)
                                        ################################# MORRIS SAMPLING###################################
                problem= {'num_vars': 17,
                              'names':['Q_90_yr','CU_yr_gw','livestock_calories','CU_yr_Surf','domestic','Agri1','SUM_GWR_K',
                                       'Fisheries','malnutrition','Dom3','Dom1','Dom2','Pi','agri_livestock','WR_fisheries','MEAN_FG_','WR_agriculture'],
                            'groups': None,
                             'bounds':[[parameters_deterministic_values['Q_90_yr']*deltas['Q_90_yr']['min'], parameters_deterministic_values['Q_90_yr']*deltas['Q_90_yr']['max']],[parameters_deterministic_values['CU_yr_gw']*deltas['CU_yr_gw']['min'], parameters_deterministic_values['CU_yr_gw']*deltas['CU_yr_gw']['max']],loc_scale(parameters_deterministic_values['livestock_calories'], parameters_uncertainty_values['livestock_calories']),
                                        [parameters_deterministic_values['CU_yr_Surf']*deltas['CU_yr_Surf']['min'], parameters_deterministic_values['CU_yr_Surf']*deltas['CU_yr_Surf']['max']],[parameters_deterministic_values['domestic']*deltas['domestic']['min'], parameters_deterministic_values['domestic']*deltas['domestic']['max']],loc_scale(parameters_deterministic_values['Agri1'], parameters_uncertainty_values['Agri1']),
                                         [parameters_deterministic_values['SUM_GWR_K']*deltas['SUM_GWR_K']['min'], parameters_deterministic_values['SUM_GWR_K']*deltas['SUM_GWR_K']['max']],loc_scale(parameters_deterministic_values['Fisheries'], parameters_uncertainty_values['Fisheries']),[parameters_deterministic_values['malnutrition']*deltas['malnutrition']['min'], parameters_deterministic_values['malnutrition']*deltas['malnutrition']['max']],
                                          loc_scale(parameters_deterministic_values['Dom3'], parameters_uncertainty_values['Dom3']),loc_scale(parameters_deterministic_values['Dom1'], parameters_uncertainty_values['Dom1']),loc_scale(parameters_deterministic_values['Dom2'], parameters_uncertainty_values['Dom2']),loc_scale(parameters_deterministic_values['Pi'], parameters_uncertainty_values['Pi']),
                                            loc_scale(parameters_deterministic_values['agri_livestock'], parameters_uncertainty_values['agri_livestock']),loc_scale(parameters_deterministic_values['WR_fisheries'], parameters_uncertainty_values['WR_fisheries']),loc_scale(parameters_deterministic_values['MEAN_FG_'], parameters_uncertainty_values['MEAN_FG_']),
                                             loc_scale(parameters_deterministic_values['WR_agriculture'], parameters_uncertainty_values['WR_agriculture'])]
                                            }
                param_vals_dists_code = morris_sampling.sample(problem, 50, num_levels=4, grid_jump=2)
                                        ###################################### FIN MORRIS ################################
                                        ######################################## SOBOL SAMPLING###########################################   
                prob_dists_code= {'num_vars': 17,
                              'names':['Q_90_yr','CU_yr_gw','livestock_calories','CU_yr_Surf','domestic','Agri1','SUM_GWR_K',
                                       'Fisheries','malnutrition','Dom3','Dom1','Dom2','Pi','agri_livestock','WR_fisheries','MEAN_FG_','WR_agriculture'],
                             'bounds':[[parameters_deterministic_values['Q_90_yr'], parameters_uncertainty_values['Q_90_yr']['GSD2']],[parameters_deterministic_values['CU_yr_gw'],parameters_uncertainty_values['CU_yr_gw']['GSD2']],unif_dist_parameters(parameters_deterministic_values['livestock_calories'], parameters_uncertainty_values['livestock_calories']),
                                        [parameters_deterministic_values['CU_yr_Surf'],parameters_uncertainty_values['CU_yr_Surf']['GSD2']],[parameters_deterministic_values['domestic'], parameters_uncertainty_values['domestic']['GSD2']],unif_dist_parameters(parameters_deterministic_values['Agri1'], parameters_uncertainty_values['Agri1']),[parameters_deterministic_values['SUM_GWR_K'], parameters_uncertainty_values['SUM_GWR_K']['GSD2']],
                                          unif_dist_parameters(parameters_deterministic_values['Fisheries'], parameters_uncertainty_values['Fisheries']),[parameters_deterministic_values['malnutrition'],parameters_uncertainty_values['malnutrition']['GSD2']],unif_dist_parameters(parameters_deterministic_values['Dom3'], parameters_uncertainty_values['Dom3']),
                                           unif_dist_parameters(parameters_deterministic_values['Dom1'], parameters_uncertainty_values['Dom1']),unif_dist_parameters(parameters_deterministic_values['Dom2'], parameters_uncertainty_values['Dom2']),unif_dist_parameters(parameters_deterministic_values['Pi'], parameters_uncertainty_values['Pi']),
                                            unif_dist_parameters(parameters_deterministic_values['agri_livestock'], parameters_uncertainty_values['agri_livestock']),unif_dist_parameters(parameters_deterministic_values['WR_fisheries'], parameters_uncertainty_values['WR_fisheries']),unif_dist_parameters(parameters_deterministic_values['MEAN_FG_'], parameters_uncertainty_values['MEAN_FG_']),
                                             unif_dist_parameters(parameters_deterministic_values['WR_agriculture'], parameters_uncertainty_values['WR_agriculture'])],
                            'dists':['lognorm','lognorm','unif','lognorm','lognorm','unif','lognorm','unif','lognorm','unif','unif','unif','unif','unif','unif','unif','unif']
                                             }    
                param_vals_sobol = saltelli.sample(prob_dists_code, 100,calc_second_order=True)
                Y_dists_Sobol = np.empty([param_vals_sobol.shape[0]])
                for i in range(0,len(param_vals_sobol)):
                     sample_sobol=dict(Q_90_yr=param_vals_sobol[i][0],CU_yr_gw=param_vals_sobol[i][1],livestock_calories=param_vals_sobol[i][2],CU_yr_Surf=param_vals_sobol[i][3],domestic=param_vals_sobol[i][4],Agri1=param_vals_sobol[i][5],SUM_GWR_K=param_vals_sobol[i][6],
                                       Fisheries=param_vals_sobol[i][7],malnutrition=param_vals_sobol[i][8],Dom3=param_vals_sobol[i][9],Dom1=param_vals_sobol[i][10],Dom2=param_vals_sobol[i][11],Pi=param_vals_sobol[i][12],agri_livestock=param_vals_sobol[i][13],
                                       WR_fisheries=param_vals_sobol[i][14],MEAN_FG_=param_vals_sobol[i][15],WR_agriculture=param_vals_sobol[i][16])
                     selected_sobol_sample=select_SA_sample(sample_sobol, parameters_deterministic_values)
                     CF_sobol, alpha, alpha_star = calculate_CF(selected_sobol_sample)
                     Y_dists_Sobol[i]=CF_sobol
                        #################################### FIN SOBOL###########################################  
#                if  'NA' not in CFs and sum(CFs) > 0. :
                if  'NA' not in CFs and deterministic_CF > 0. :
                    
                    ################################################## SOBOL ANALYSIS ##################################
                    Si= sobol.analyze(prob_dists_code,Y_dists_Sobol,calc_second_order=True,print_to_console=False)

                    sobol_indices={}
                    sobol_indices['S1']={}
                    sobol_indices['ST']={}
                    for i in prob_dists_code['names']:
                        sobol_indices['S1'][i]=Si['S1'][prob_dists_code['names'].index(i)]
                        sobol_indices['ST'][i]=Si['ST'][prob_dists_code['names'].index(i)]

                        to_add[0].update({'parameter': i,'S1' :sobol_indices['S1'][i],
                                           'ST' :sobol_indices['ST'][i]})
                        Sobol_result = concat([Sobol_result, DataFrame(to_add).transpose()])
                    for i in range(len(CFs)):
                        if 0. in CFs:
                            CFs.remove(0.)
                    if len(CFs)>200. :
                        CF_sobol_ratio.append((np.mean(CFs)/np.mean(Y_dists_Sobol))*100.)
                    print CF_sobol_ratio
                    to_add[0].update({'CF sobol ratio':CF_sobol_ratio})
                    ############################################# END OF SOBOL ANALYSIS #######################################
                ########################################################### MORRIS ANALYSIS#############################################################################                
                    Y_dists_code = np.empty([param_vals_dists_code.shape[0]])
                    for i in range(0,len(param_vals_dists_code)):
                         morris_sample=dict(Q_90_yr=param_vals_dists_code[i][0],CU_yr_gw=param_vals_dists_code[i][1],livestock_calories=param_vals_dists_code[i][2],CU_yr_Surf=param_vals_dists_code[i][3],domestic=param_vals_dists_code[i][4],Agri1=param_vals_dists_code[i][5],SUM_GWR_K=param_vals_dists_code[i][6],
                                           Fisheries=param_vals_dists_code[i][7],malnutrition=param_vals_dists_code[i][8],Dom3=param_vals_dists_code[i][9],Dom1=param_vals_dists_code[i][10],Dom2=param_vals_dists_code[i][11],Pi=param_vals_dists_code[i][12],agri_livestock=param_vals_dists_code[i][13],
                                           WR_fisheries=param_vals_dists_code[i][14],MEAN_FG_=param_vals_dists_code[i][15],WR_agriculture=param_vals_dists_code[i][16])
                         Ishigami_selected_sample=select_SA_sample(morris_sample, parameters_deterministic_values)
                         CF_Morris, alpha, alpha_star = calculate_CF(Ishigami_selected_sample)
                         Y_dists_code[i]=CF_Morris
                    MSI = morris.analyze(problem, param_vals_dists_code, Y_dists_code, conf_level=0.95,print_to_console=False, num_levels=4, grid_jump=2)
                    if sum(Y_dists_code)>0.:
                        CF_Morris_ratio.append((np.mean(CFs)/np.mean(Y_dists_code))*100.)
                    print CF_Morris_ratio
                    to_add[0].update({'CF Morris ratio': CF_Morris_ratio})
                    fig = plt.figure()
                    ax = fig.add_subplot(111)
                    Morris_indices={}
                    Morris_indices['mu_star']={}
                    Morris_indices['sigma']={}
                    for i in problem['names']:
                        Morris_indices['mu_star'][i]=MSI['mu_star'][problem['names'].index(i)]
                        Morris_indices['sigma'][i]=MSI['sigma'][problem['names'].index(i)]
                        to_add[0].update({'parameter': i,'mu_star' :Morris_indices['mu_star'][i],
                                           'sigma' :Morris_indices['sigma'][i] })
                        Morris_result = concat([Morris_result, DataFrame(to_add).transpose()])

                    param_dict={}
#                p = morris_plotting.covariance_plot(ax,Morris_ind,param_dict, unit="")
                    Horiz_bar_plot = morris_plotting.horizontal_bar_plot(ax,MSI,param_dict,sortby='mu_star', unit="")
                    ########################################################## END OF THE MORRIS ANALYSIS#############################################################################
                    if deterministic_CF!= 0.:
                        to_add[0].update({'CF deterministic': deterministic_CF})
                    else:
                        to_add[0].update({'CF deterministic': 0.})
                elif 'NA' in CFs:
                    to_add[0].update({'CF deterministic': 'division par 0(water availability= 0)'})  
                elif sum(CFs)== 0.:
                    to_add[0].update({'CF deterministic': deterministic_CF})                                       
            else:
                to_add[0].update({'CF deterministic': 'missing data'})
            Sobol_result = concat([Sobol_result, DataFrame(to_add).transpose()])
            Morris_result = concat([Morris_result, DataFrame(to_add).transpose()])


cols = ['cell_ID', 'country', 'ground_surface', 'water_type']
cols.extend(['CF deterministic'])
cols.extend(['CF sobol ratio'])
cols.extend(['parameter','S1','ST'])
filename = 'sobol_result.xlsx'
folder = '/Users/Omar/Desktop/Projet_recherche/Water Use/Sensitivity analysis/Code'
Sobol_result.to_excel(os.path.join(folder, filename), columns = cols, merge_cells = False)

#cols = ['cell_ID', 'country', 'ground_surface', 'water_type','CF deterministic','CF Morris ratio']
#cols.extend(['parameter','mu_star','sigma'])
#filename = 'Morris_result.csv'
#folder = '/Users/Omar/Desktop/Projet_recherche/Water Use/Sensitivity analysis/Code'
#Morris_result.to_csv(os.path.join(folder, filename), columns = cols, merge_cells = False)