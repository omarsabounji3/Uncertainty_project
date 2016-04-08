#code pour l'analyse de sensibilite des seuils min et max
from openpyxl.reader.excel import load_workbook
import numpy as np
from scipy.stats import lognorm, norm, beta, uniform, anderson,triang
from scipy.stats.mstats import spearmanr
import os
import scipy as sp
from cPickle import load, dump
from pandas import read_excel, DataFrame, concat
from copy import deepcopy
import random as rnd 
import matplotlib.pyplot as mpl
import matplotlib.pyplot as plt
from matplotlib.pyplot import title, xlabel, ylabel
from matplotlib.backends.backend_pdf import PdfPages
import statsmodels.api as sm
from numpy import sort, linspace, log
from matplotlib.pyplot import figure, axis, show
from scipy.stats import norm, kstest
from statsmodels.sandbox.regression.predstd import wls_prediction_std
from numpy.random import dirichlet

# Fonction qui permet de lire le fichier excel

def read_data(filename):

    print 'loading ' + filename + ', takes a few seconds'

    wb = load_workbook(filename)


    ws = wb.get_sheet_by_name('Cell data')
    
    print( 'reading cell data')

    cell_data = {}

    for cell_number in range(1, 809):

        cell_data[cell_number] = {}

    row_number = 0

    header = {}

    for row in ws.rows:
    #for row in []:

        row = list(row)

        row_number += 1

        if [1, 2, 3].count(row_number) == 1:

            for column in range(len(row)):

                try:

                    row[column] = str(row[column].value)

                except TypeError:

                    row[column] = ''

            header[row_number] = row

        else:

            cell_number = float(row[0].value)

            country = str(row[1].value)

            cell_data[cell_number]['Country'] = str(row[1].value)

            for column in range(2, 7):

                tag = header[3][column]

                cell_data[cell_number][tag] = float(row[column].value)

            for column in range(7, len(row)):

                tag1 = header[1][column]

                tag2 = header[2][column]

                tag3 = header[3][column]

                try:

                    cell_data[cell_number][tag1]

                except KeyError: 

                    cell_data[cell_number][tag1] = {}

                try:

                    cell_data[cell_number][tag1][tag2]

                except KeyError: 

                    cell_data[cell_number][tag1][tag2] = {}

                try:

                    element = float(row[column].value)

                except ValueError:

                    element = str(row[column].value)
                    
                except TypeError:

                    element = str(row[column].value)
                
                cell_data[cell_number][tag1][tag2][tag3] = element

    

    ws = wb.get_sheet_by_name('Country data')
    print ('reading Country data')
    country_data = {}

    row_number = 0

    for row in ws.rows:
    #for row in []:

        row = list(row)

        row_number += 1

        if row_number != 1:

            country = str(row[0].value)

            try:

                AC = float(row[1].value)

            except ValueError:

                AC = 'NA'

            country_data[country] = {}

            country_data[country]['AC'] = AC


    

    ws = wb.get_sheet_by_name('User distribution')
    print ('reading User distribution')
    user_dist = {}

    row_number = 0

    header = {}

    for row in ws.rows:
    #for row in []:
        
        row = list(row
        )

        row_number += 1

        if [1, 2, 3].count(row_number) == 1:

            for column in range(len(row)):

                try:

                    row[column] = str(row[column].value)

                except TypeError:

                    row[column] = ''

            header[row_number] = row

        else:

            cell_number = int(row[0].value)

            user_dist[cell_number] = {}

            for column in range(1,len(row)):

                ground_surface = header[1][column]

                try:

                    user_dist[cell_number][ground_surface]

                except KeyError:

                    user_dist[cell_number][ground_surface] = {}

                water_type = header[2][column]

                try:

                    user_dist[cell_number][ground_surface][water_type]

                except KeyError:

                    user_dist[cell_number][ground_surface][water_type] = {}

                user = header[3][column]

                if user != 'Hydro' and user != 'Recre':

                    try:

                        data = float(row[column].value)

                    except TypeError:

                        data = 'NA'

                    user_dist[cell_number][ground_surface][water_type][user] = data
                    
    ws = wb.get_sheet_by_name('Uncertainty data')
    print ('reading uncertainty data')
    uncertainty_data = {}
    
    for row in ws.rows:
    #for row in []:
        
        variable = row[0]
        data_type = row[1]
        data = row[2]
    
        variable = str(variable.value)
    
        data_type = str(data_type.value)
        
        if data.value == None :  
            break
        try:
            data = float(data.value)
        except ValueError:
            
            data= str(data.value)
    
        try:
        
        
            uncertainty_data[variable]
    
        except KeyError:
        
            uncertainty_data[variable] = {}
            
        uncertainty_data[variable][data_type] = data
        
    ws = wb.get_sheet_by_name('Equation data')
    print( 'reading equation data')
    eq_data = {}
        
    for row in ws.rows:
        
        variable= row[0]
        
        data=row[1]
        
        variable = str(variable.value)
        
        data = float(data.value)
        
        eq_data[variable] = data
    
    ws = wb.get_sheet_by_name('EF')
    print ('reading EF')
    EF_data = {}
    for row in ws.rows:
    #for row in []:
        
        variable = row[0]
        data_type = row[1]
        data = row[2]
        unit = row[3]
        
        variable = str(variable.value)
        
        data_type = str(data_type.value)
        
        data = float(data.value)
        try:
            
            EF_data[variable]
        except KeyError:
            
            EF_data[variable] = {}
            
        EF_data[variable][data_type] = data

    return cell_data, country_data, user_dist, eq_data, uncertainty_data, EF_data
def save_variable(all_info):
    #all_info = [
    #           [path, variable_name, variable], 
    #           [path, variable_name, variable]
    #           ]
    for path, variable_name, variable in all_info:
        filename = os.path.join(path, variable_name + '.pkl')
        file = open(filename, 'wb')
        dump(variable, file)
        file.close()
def load_variables(all_info):
    #all_info = [[path, variable_name], [path, variable_name], ...]
    variables = []
    for path, variable_name in all_info:
        filename = path + '/' + variable_name + '.pkl'
        variables.append(load(open(filename, 'rb')))
    return variables
def build_file_list(dirpath, extension = None):
    pre_list = os.listdir(dirpath)
    if extension == None:
        filelist = [filename for filename in pre_list 
                if os.path.isfile(os.path.join(dirpath, filename)) and '~' not in filename]
    else:
        filelist = [filename for filename in pre_list 
                if os.path.isfile(os.path.join(dirpath, filename))
                and filename.split('.')[-1].lower() == extension and '~' not in filename]
    
    return filelist

# Fonction de calcul du CF 

def calculate_CF(parameters):
    users_aggregation = {}
    users_aggregation['WR_agriculture'] = ['Agri1']
    users_aggregation['domestic'] = ['Dom1', 'Dom2', 'Dom3']
    users_aggregation['WR_fisheries'] = ['Fisheries']
    deterministic_CF = 0.
    alpha = 'NA'
    alpha_star= 'NA'
    x3=[]
    if parameters['AC'] != 1.:
        if ((parameters['SUM_GWR_K'] == 0. and ground_surface == 'groundwater') or 
                (parameters['Q_90_yr'] == 0. and ground_surface == 'surface')):
            deterministic_CF = 'NA'
        else:
            if ground_surface == 'groundwater':
                alpha_star = parameters['CU_yr_gw']* parameters['MEAN_FG_'] / parameters['SUM_GWR_K'] / parameters['Pi']
            else:
                alpha_star = parameters['CU_yr_Surf']*(1.-parameters['MEAN_FG_'])/ parameters['Q_90_yr'] / parameters['Pi']
#            print alpha_star
            if alpha_star < parameters['min']:
                alpha = 0.
            elif alpha_star > parameters['max']:
                alpha = 1.
            else:
                #Equation du alpha star
                alpha = (alpha_star-parameters['min'])/(parameters['max'] - parameters['min'])
                #alpha = eq_data['K'] / ((1. + (eq_data['Q'] * exp(-eq_data['B'] * (alpha_star - eq_data['M'])))**(1. / eq_data['v'])))
            if alpha == 0. :
                deterministic_CF= 0.
            else:
                for user_general in users_aggregation:
    #                print user_general
                    if user_general == 'domestic':
                        EF = parameters[user_general]
                        x1=parameters[user_general]*parameters[user_general]
                    elif user_general == 'WR_fisheries':
                        EF = parameters['malnutrition'] / parameters[user_general]
                    else:
                        EF_uncorrected = parameters['malnutrition'] / parameters[user_general]
                        EF = (((1. - parameters['agri_livestock']) * EF_uncorrected) + 
                                (parameters['agri_livestock']* EF_uncorrected / parameters['livestock_calories']))

                    for user_specific in users_aggregation[user_general]:
                        deterministic_CF += alpha * (1. - parameters['AC']) * EF * parameters[user_specific]
    return deterministic_CF, alpha, alpha_star


def create_rv(parameters_deterministic_values, parameters_uncertainty_values):
    RVs = {}
    all_data_available= True
    for parameter in parameters_uncertainty_values:
        if parameters_deterministic_values[parameter]==0.:
            RVs[parameter]= uniform(loc = 0., scale = 0.) 
        elif parameters_uncertainty_values[parameter]['distribution'] == 'uniform':
            if 'width' in parameters_uncertainty_values[parameter]:
                loc = parameters_deterministic_values[parameter] - parameters_uncertainty_values[parameter]['width']
                if loc< 0.:
                    1/0
                scale = 2.*parameters_uncertainty_values[parameter]['width']
            else:
                loc= parameters_uncertainty_values[parameter]['min']
                scale = parameters_uncertainty_values[parameter]['max'] -parameters_uncertainty_values[parameter]['min']
            RVs[parameter] = uniform(loc = loc, scale = scale) 
        elif parameters_uncertainty_values[parameter]['distribution'] == 'normal':
            RVs[parameter] = norm(loc = parameters_deterministic_values[parameter] , scale = parameters_uncertainty_values[parameter]['std'])
        elif parameters_uncertainty_values[parameter]['distribution'] == 'lognormal':
            RVs[parameter] = fit_lognormal(parameters_deterministic_values[parameter], parameters_uncertainty_values[parameter]['GSD2'])
        elif parameters_uncertainty_values[parameter]['distribution'] == 'beta':
#            RVs[parameter] = create_beta(parameters_deterministic_values[parameter], parameters_uncertainty_values[parameter])
            if 'width' in parameters_uncertainty_values[parameter]:
                loc = parameters_deterministic_values[parameter] - parameters_uncertainty_values[parameter]['width']
                if loc< 0.:
                    loc=0.
                scale = 2.*parameters_uncertainty_values[parameter]['width']
            else:
                loc= parameters_uncertainty_values[parameter]['min']
                scale = parameters_uncertainty_values[parameter]['max'] -parameters_uncertainty_values[parameter]['min']
            RVs[parameter] = uniform(loc = loc, scale = scale) 
    return RVs, all_data_available
# Fonction qui diagnostique la zone ou se trouve l'alpha star
def alpha_star_analysis(parameters_uncertainty_values, min_alpha_stars, max_alpha_stars):
    contents = []
    if parameters_uncertainty_values['min']['min'] > min_alpha_stars and parameters_uncertainty_values['min']['min'] < max_alpha_stars:
        contents.append(1)
    if parameters_uncertainty_values['min']['max'] > min_alpha_stars and parameters_uncertainty_values['min']['max'] < max_alpha_stars:
        contents.append(2)
    if parameters_uncertainty_values['max']['min'] > min_alpha_stars and parameters_uncertainty_values['max']['min'] < max_alpha_stars:
        contents.append(3)
    if parameters_uncertainty_values['max']['max'] > min_alpha_stars and parameters_uncertainty_values['max']['max'] < max_alpha_stars:
        contents.append(4)
    return contents
def fit_lognormal(average, GSD2):
    s = np.log(GSD2**.5)
    rv = lognorm(s, scale = average)
    rv.ppf([.0225, .5, .9775])
    return rv
def create_beta(average, parameters_uncertainty_values):
    if 'width' in parameters_uncertainty_values:
        minimum = average - parameters_uncertainty_values['width']
        maximum = average + parameters_uncertainty_values['width']
    else:
        minimum = parameters_uncertainty_values['min']
        maximum = parameters_uncertainty_values['max']
    if minimum < 0.:
        minimum =0.
    p975 = minimum + (maximum - minimum)*.975
    assert maximum > minimum, 'minimum should be lower than maximum'
    assert average > minimum, 'average should be higher than minimum'
    assert average < maximum, 'average should be lower than maximum'
    assert p975 > average, 'p975 should be higher than average'
    mu = (average - minimum) / (maximum - minimum)
    ba_ratio = (1 - mu)/mu
    scale = maximum - minimum
    #initial guess
    a = 1.
    b = a*ba_ratio
    threshold = .01
    direction = 0
    delta = .2
    counter = 0
    while 1:
        counter += 1
        rv = beta(a, b, loc = minimum, scale = scale)
        if abs(rv.ppf(.975)/p975) < 1.+threshold and abs(rv.ppf(.975)/p975 > 1.-threshold):
            break
        elif rv.ppf(.975) > p975:
            if direction == -1:
                delta = delta/2.
            direction = 1
        else:
            if direction == 1:
                delta = delta/2.
            direction = -1
        if counter > 1000:
            print (counter)
        a += direction*delta
        b = a*ba_ratio
    return rv
def create_sample(RVs, nb_iteration):
 #create a sample of nb_iteration for each parameter
    sample = {}
    for p in RVs:
        sample[p] = RVs[p].rvs(nb_iteration)
    return sample
def select_sample(sample, all_data, n):
    selected_data= {}
    for p in parameters_deterministic_values:
        if p in sample:
            selected_data[p]= sample[p][n]
        else:
            selected_data[p]=parameters_deterministic_values[p]
    return selected_data


def plot_histogram(data, variable_name, mean, deterministic, pdf_files, Filename,nb_iteration):
    fig = mpl.figure(1)
    ax1 = fig.add_subplot(111)
#    ax1.set_xlim(min_alphas,max_alphas*0.975)
    ax1.hist(data, 50., facecolor='grey', alpha=0.9)
    plt.axvline(x=mean, ymin=0.0, ymax = 1, linewidth=2, color='r')
    plt.axvline(x=deterministic, ymin=0.0, ymax = 1, linewidth=2, color='b')
    plt.legend(['deterministic', 'MC average'])
    ylabel('Count')
    xlabel(variable_name)
    title(Filename.replace('.pdf', ', %s iter.'% nb_iteration))
    pdf_files.savefig(fig)
    if max(data)/np.percentile(data, 97.5)> 2:
        fig = mpl.figure(1)
        ax1 = fig.add_subplot(111)
        ax1.set_xlim(min(data),np.percentile(data, 97.5))
        ax1.hist(data, 50., facecolor='purple', alpha=0.9)
        plt.axvline(x=mean, ymin=0.0, ymax = 1, linewidth=2, color='r')
        plt.axvline(x=deterministic, ymin=0.0, ymax = 1, linewidth=2, color='b')
        plt.legend(['deterministic', 'MC average'])
        ylabel('Count')
        xlabel(variable_name)
        title(Filename.replace('.pdf', ', %s iter, rescaled'% nb_iteration))
        pdf_files.savefig(fig)  
    mpl.show()
    return pdf_files
def stats_on_CF(CFs, alphas, alpha_stars):
    m = np.mean(CFs)
    m_alpha = np.mean(alphas)
    m_alpha_star=np.mean(alpha_stars)
    m = np.mean(CFs)
    mini=np.min(CFs)
    maxi=np.max(CFs)
    confidence_95 = [np.percentile(CFs, 2.5), np.percentile(CFs, 97.5)]
    median = np.percentile(CFs, 50.)
    min_alphas=np.min(alphas)
    min_alpha_stars=np.min(alpha_stars)
    max_alphas=np.max(alphas)
    max_alpha_stars=np.max(alpha_stars)
    return m, mini, maxi, confidence_95, median, min_alpha_stars, max_alpha_stars, min_alphas, max_alphas, m_alpha,m_alpha_star 
def extract_data(cell_ID, ground_surface, water_type, cell_data, country_data, eq_data,EF_data):
    all_data_available = False
    while 1:
        parameters = {}
        if country_data[cell_data[cell_ID]['Country']]['AC'] == 'NA':
            break
        parameters['AC'] = country_data[cell_data[cell_ID]['Country']]['AC']
        parameters['CU_yr_gw'] = cell_data[cell_ID]['CU_yr_gw']
        parameters['SUM_GWR_K'] = cell_data[cell_ID]['SUM_GWR_K']
        parameters['CU_yr_Surf'] = cell_data[cell_ID]['CU_yr_Surf']
        parameters['Q_90_yr'] = cell_data[cell_ID]['Q_90_yr']
        parameters['MEAN_FG_']= cell_data[cell_ID]['MEAN_FG_']
        if cell_data[cell_ID]['Pi'][ground_surface][water_type] in ['NA', '99999']:
            break
        parameters['Pi']=cell_data[cell_ID]['Pi'][ground_surface][water_type]
        parameters.update(user_dist[cell_ID][ground_surface][water_type])
        all_data_available = True
        break
    parameters['min'] = eq_data['min_1']
    parameters['max'] = eq_data['max_1']
    parameters['malnutrition']=EF_data['malnutrition']['average']
    parameters['domestic'] = EF_data['domestic']['average']
    parameters['WR_agriculture'] = EF_data['WR_agriculture']['average']
    parameters['WR_fisheries']= EF_data['WR_fisheries']['average']
    parameters['agri_livestock'] = EF_data['agri_livestock']['average']
    parameters['livestock_calories'] = EF_data['livestock_calories']['average']
    return parameters, all_data_available


def extract_uncertainty_info(cell_ID, ground_surface, water_type, cell_data, country_data, eq_data, EF_data, uncertainty_data):
	#put in a dictionary like this:

    parameters_uncertainty_values = {}
    parameters_uncertainty_values['malnutrition'] = {}
    parameters_uncertainty_values['domestic'] = {}
    parameters_uncertainty_values['CU_yr_Surf'] = {}
    parameters_uncertainty_values['CU_yr_gw'] = {}
    parameters_uncertainty_values['Q_90_yr'] = {}
    parameters_uncertainty_values['SUM_GWR_K'] = {}
    parameters_uncertainty_values['agri_livestock'] = {}
    parameters_uncertainty_values['MEAN_FG_'] = {}
    parameters_uncertainty_values['Pi'] = {}
    parameters_uncertainty_values['Fisheries'] = {}
    parameters_uncertainty_values['Dom1'] = {}
    parameters_uncertainty_values['Dom2'] = {}
    parameters_uncertainty_values['Dom3'] = {}
    parameters_uncertainty_values['Agri1'] = {}
#    parameters_uncertainty_values['Ind'] = {}
#    parameters_uncertainty_values['Cooling'] = {}
    parameters_uncertainty_values['min'] = {}
    parameters_uncertainty_values['max'] = {}
    parameters_uncertainty_values['WR_fisheries'] = {}
    parameters_uncertainty_values['WR_agriculture'] = {}
    parameters_uncertainty_values['livestock_calories'] = {}
    
    
    parameters_uncertainty_values['malnutrition']['GSD2'] = EF_data['malnutrition']['GSD2']
    parameters_uncertainty_values['domestic']['GSD2']=EF_data['domestic']['GSD2']
    parameters_uncertainty_values['CU_yr_Surf']['GSD2']=uncertainty_data['CU_yr_Surf']['GSD2']
    parameters_uncertainty_values['CU_yr_gw']['GSD2']=uncertainty_data['CU_yr_gw']['GSD2']
    parameters_uncertainty_values['Q_90_yr']['GSD2']=uncertainty_data['Q_90_yr']['GSD2']
    parameters_uncertainty_values['SUM_GWR_K']['GSD2']=uncertainty_data['SUM_GWR_K']['GSD2']

    parameters_uncertainty_values['agri_livestock']['width']=EF_data['agri_livestock']['width']
    parameters_uncertainty_values['MEAN_FG_']['width']=uncertainty_data['MEAN_FG_']['width']
    parameters_uncertainty_values['Pi']['width']=uncertainty_data['Pi']['width']
    parameters_uncertainty_values['Fisheries']['width']=uncertainty_data['user_distribution']['width']
    parameters_uncertainty_values['Dom1']['width']=uncertainty_data['user_distribution']['width']
    parameters_uncertainty_values['Dom2']['width']=uncertainty_data['user_distribution']['width']
    parameters_uncertainty_values['Dom3']['width']=uncertainty_data['user_distribution']['width']
    parameters_uncertainty_values['Agri1']['width']=uncertainty_data['user_distribution']['width']    
#    parameters_uncertainty_values['Ind']['width']=uncertainty_data['user_distribution']['width']
#    parameters_uncertainty_values['Cooling']['width']=uncertainty_data['user_distribution']['width']
    parameters_uncertainty_values['min']['min']=eq_data['min_1_low']
    parameters_uncertainty_values['min']['max']=eq_data['min_1_high']
    parameters_uncertainty_values['max']['min']=eq_data['max_1_low']
    parameters_uncertainty_values['max']['max']=eq_data['max_1_high']
    parameters_uncertainty_values['WR_fisheries']['min']=EF_data['WR_fisheries']['min']
    parameters_uncertainty_values['WR_fisheries']['max']=EF_data['WR_fisheries']['max']
    parameters_uncertainty_values['WR_agriculture']['min']=EF_data['WR_agriculture']['min']
    parameters_uncertainty_values['WR_agriculture']['max']=EF_data['WR_agriculture']['max']
    parameters_uncertainty_values['livestock_calories']['min']=EF_data['livestock_calories']['min']
    parameters_uncertainty_values['livestock_calories']['max']=EF_data['livestock_calories']['max']
    
    
    parameters_uncertainty_values['Pi']['distribution']=uncertainty_data['Pi']['distribution']
    parameters_uncertainty_values['MEAN_FG_']['distribution']=uncertainty_data['MEAN_FG_']['distribution']
    parameters_uncertainty_values['SUM_GWR_K']['distribution']=uncertainty_data['SUM_GWR_K']['distribution']
    parameters_uncertainty_values['Q_90_yr']['distribution']=uncertainty_data['Q_90_yr']['distribution']
    parameters_uncertainty_values['CU_yr_gw']['distribution']=uncertainty_data['CU_yr_gw']['distribution']
    parameters_uncertainty_values['CU_yr_Surf']['distribution']=uncertainty_data['CU_yr_Surf']['distribution']
    parameters_uncertainty_values['livestock_calories']['distribution']=uncertainty_data['livestock_calories']['distribution']
    parameters_uncertainty_values['agri_livestock']['distribution']=uncertainty_data['agri_livestock']['distribution']
    parameters_uncertainty_values['domestic']['distribution']=uncertainty_data['domestic']['distribution']
    parameters_uncertainty_values['WR_agriculture']['distribution']= uncertainty_data['WR_agriculture']['distribution']
    parameters_uncertainty_values['WR_fisheries']['distribution']=uncertainty_data['WR_fisheries']['distribution']
    parameters_uncertainty_values['malnutrition']['distribution']=uncertainty_data['malnutrition']['distribution']
    parameters_uncertainty_values['Fisheries']['distribution']=uncertainty_data['Fisheries']['distribution']
    parameters_uncertainty_values['Dom1']['distribution']=uncertainty_data['Dom1']['distribution']
    parameters_uncertainty_values['Dom2']['distribution']=uncertainty_data['Dom2']['distribution']
    parameters_uncertainty_values['Dom3']['distribution']=uncertainty_data['Dom3']['distribution']
    parameters_uncertainty_values['Agri1']['distribution']=uncertainty_data['Agri1']['distribution']    
#    parameters_uncertainty_values['Ind']['distribution']=uncertainty_data['Ind']['distribution']
#    parameters_uncertainty_values['Cooling']['distribution']=uncertainty_data['Cooling']['distribution']
    parameters_uncertainty_values['min']['distribution']=uncertainty_data['min']['distribution']
    parameters_uncertainty_values['max']['distribution']=uncertainty_data['max']['distribution']    
    
    return parameters_uncertainty_values


def test_anderson(sample,to_add, df):
    sample = list(sample)
    nb_0 =sample.count(0.)
    sample = filter(lambda a: a != 0., sample)
    anderson_limits = [.15, .1, .05, .025, .01]
    A2, critical, sig = anderson(np.log10(sample))
    if sum(A2 < critical) == 0:
        anderson_status = 'rejected'
    else:
        anderson_status = anderson_limits[sum(A2 < critical) - 1]
        sample_normal=norm.fit(sample)
    to_add[0].update({'nombre de zero':nb_0,'anderson status':anderson_status})
    if anderson_status != 'rejected' :
        to_add[0].update({'normal average':sample_normal[0], 'normal std':sample_normal[1]})
    df=concat([df, DataFrame(to_add).transpose()])
    
    return df


def perform_rank_correlation(sample, CFs, alphas, alpha_stars, spearman_result, to_add): 
    #print 'spearman'
    p = ['Q_90_yr', 'CU_yr_gw', 'CU_yr_Surf', 'SUM_GWR_K', 'Pi', 'MEAN_FG_']
    if len(set(alphas)) == 1:
        to_add[0].update({'alpha': 0.})
        to_add[0].update({'alpha star': 0.})
    else:
        to_add[0].update({'alpha': spearmanr(alphas, CFs)[0]})
        to_add[0].update({'alpha star': spearmanr(alpha_stars, CFs)[0]})
    for parameter in sample:
        if len(set(alphas)) == 1 and parameter in p:
            to_add[0].update({parameter: 0.})
        else:
            to_add[0].update({parameter: spearmanr(sample[parameter], CFs)[0]})
    spearman_result = concat([spearman_result, DataFrame(to_add).transpose()])
    #if to_add[0]['cell_ID'] == 3 and to_add[0]['water_type'] == '2d' and to_add[0]['ground_surface'] == 'surface':
    #    1/0
    return spearman_result
folder = '/Users/Omar/Desktop/Projet_recherche/Water Use/Sensitivity analysis/Code'
#folder = r'C:\python\water_use_MC_2\src'
force_read_excel=False

def reg_m(y, x):
    ones = np.ones(len(x[0]))
    X = sm.add_constant(np.column_stack((x[0], ones)))
    for ele in x[1:]:
        X = sm.add_constant(np.column_stack((ele, X)))
    results = sm.OLS(y, X).fit()
    return results

def test_kolmogorov(sample):
    
    test_result = {}
    #test for lognormal
    sample_log = log(sample)
    parameters = norm.fit(sample_log)
    (test_stats_log, p_value_log) = kstest(sample_log, 'norm', parameters)
    test_result['lognorm'] = {}
    test_result['lognorm']['parameters'] = parameters
    test_result['lognorm']['Kolmogorov test'] = (test_stats_log, p_value_log)
    # Interpreter directement les parametres utilisable pour une lognormale
    if test_stats_log > p_value_log:
        test_result['lognorm']['status'] = 'rejected'
    else:
        test_result['lognorm']['status'] = 'accepted'
    
    #test for beta
#    normalized_sample = sample / max(sample)
    parameters = beta.fit(sample)
    (test_stats, p_value) = kstest(sample, 'beta', parameters)
    test_result['beta standard'] = {}
    test_result['beta standard']['parameters'] = parameters
    test_result['beta standard']['Kolmogorov test'] = (test_stats, p_value)
    if test_stats > p_value:
        test_result['beta standard']['status'] = 'rejected'
    else:
        test_result['beta standard']['status'] = 'accepted'
    
    # test for triangular
    
    sample_mode = (sp.stats.mode(sample)[0][0]-np.min(sample))/(np.max(sample)-np.min(sample))
    parameters = triang.fit(sample,c=sample_mode, loc=np.min(sample),scale=(np.max(sample)/sp.stats.mode(sample)[0][0]))
    (test_stats_tri, p_value_tri) = kstest(sample, 'triang', parameters)
    test_result['triangular'] = {}
    test_result['triangular']['parameters'] = parameters
    test_result['triangular']['Kolmogorov test'] = (test_stats_tri, p_value_tri)
    if test_stats_tri > p_value_tri:
        test_result['triangular']['status'] = 'rejected'
    else:
        test_result['triangular']['status'] = 'accepted'
        
#    test pour l'uniform
    parameters = uniform.fit(sample)
    (test_stats_uni, p_value_uni) = kstest(sample, 'uniform',parameters)
    test_result['uniform'] = {}
    test_result['uniform']['parameters'] = parameters
    test_result['uniform']['Kolmogorov test'] = (test_stats_uni, p_value_uni)
    if test_stats_uni > p_value_uni:
        test_result['uniform']['status'] = 'rejected'
    else:
        test_result['uniform']['status'] = 'accepted'

    return test_result
def dirichlet_sample(sample, parameters_deterministic_values):
    det_shares=[]
    sample_ud={}
    User={}
    UD={}
    sample_used={}
    User['Ind']=parameters_deterministic_values['Ind']
    User['Cooling']=parameters_deterministic_values['Cooling']
    User['Dom1']=parameters_deterministic_values['Dom1']
    User['Dom2']=parameters_deterministic_values['Dom2']
    User['Dom3']=parameters_deterministic_values['Dom3']
    User['Agri1']=parameters_deterministic_values['Agri1']

    for i in User:
        if User[i]!=0. :
            UD[i]=User[i]
    for i in UD:
        if i in ['Dom1','Dom2','Dom3','Agri1']:
            del sample[i]
            sample_used[i]=UD[i]
        det_shares.append(UD[i]) 

    scaling =[1100, 100]
    det_shares=np.array(det_shares)
    s = [dirichlet(det_shares*scaling[0], nb_iteration), 
         dirichlet(det_shares*scaling[1], nb_iteration)]
    position=[]
    det_values=list(det_shares)
    for i in sample_used:
        position.append(det_values.index(sample_used[i]))
        sample_ud[i]=[]
    for i in range(len(s[0])):
        for j in position:
            sample_ud[UD.keys()[j]].append(s[0][i][j])                
    sample.update(sample_ud)
#    if sample_used.keys() != UD.keys():
#        print 'different'
#    else :
#        print 'meme'
    width=[]
    for i in sample_ud:
        width.append(np.percentile(sample_ud[i], 97.5)-np.percentile(sample_ud[i], 0.025))
    return sample

if 'cell_data.pkl' in build_file_list(folder) and not force_read_excel :
    var_names = ['cell_data', 'country_data', 'user_dist','eq_data', 'uncertainty_data', 'EF_data']
    print ('loading pickle')
    all_info = zip([folder]*len(var_names), var_names)
    cell_data, country_data, user_dist,eq_data, uncertainty_data, EF_data = load_variables(all_info)
    print ('finish loading pickle' )
else:
    filename = 'water_use_all_info.xlsx'
    cell_data, country_data, user_dist,eq_data, uncertainty_data, EF_data = read_data(filename)
    all_info = [[folder, 'cell_data', cell_data], 
                [folder, 'country_data', country_data], 
                [folder, 'user_dist', user_dist], 
                [folder, 'eq_data', eq_data], 
                [folder, 'uncertainty_data', uncertainty_data], 
                [folder, 'EF_data', EF_data]]
    save_variable(all_info)
Kolmogorov = DataFrame()
nb_iteration = 1000
create_graph = False
anderson_result= DataFrame()
#
#for cell_ID in [1]:
#    for ground_surface in ['groundwater']:
#        for water_type in ['5']:
for cell_ID in cell_data:
    for ground_surface in cell_data[cell_ID]['Pi']:
        for water_type in cell_data[cell_ID]['Pi'][ground_surface]:
        #calculate if all data available
            print ('Cell: %s, water: %s %s' % (cell_ID, water_type, ground_surface))
            parameters_deterministic_values, all_data_available = extract_data(cell_ID, ground_surface, water_type, cell_data, country_data, eq_data,EF_data)
            to_add = {0: {'cell_ID': cell_ID, 
								'ground_surface': ground_surface, 
								'water_type': water_type, 
								'country': cell_data[cell_ID]['Country'], 
								}}
            if all_data_available:
                deterministic_CF, deterministic_alpha, deterministic_alpha_star = calculate_CF(parameters_deterministic_values)
                parameters_uncertainty_values = extract_uncertainty_info(cell_ID, ground_surface, water_type, cell_data, country_data, eq_data, EF_data, uncertainty_data)
                RVs, all_data_available = create_rv(parameters_deterministic_values, parameters_uncertainty_values)
                sample = create_sample(RVs, nb_iteration)
                sample=dirichlet_sample(sample, parameters_deterministic_values)
                CFs = []
                alpha_stars = []
                alphas = []
                for iteration in range(nb_iteration): 
                    selected_sample = select_sample(sample, parameters_deterministic_values, iteration)
                    CF, alpha, alpha_star = calculate_CF(selected_sample)
                    CFs.append(CF)
                    alpha_stars.append(alpha_star)
                    alphas.append(alpha)
#                if  'NA' not in CFs and sum(CFs) > 0.:
                if  'NA' not in CFs:
                    if sum(CFs) > 0. :
                        anderson_result = test_anderson(CFs, deepcopy(to_add),anderson_result)
#                    for i in range(len(CFs)):
#                        if 0. in CFs:
#                          CFs.remove(0.)
#                       
#                    if len(CFs)>25. :
#                    print test_kolmogorov(CFs)
                        nb_0 =CFs.count(0.)
#                        sample=CFs
#                        sample_mode = (sp.stats.mstats.mode(sample))
#                        print sample_mode
#                        print np.min(CFs)
                        test=test_kolmogorov(CFs)
                        print test['triangular']['status']
#                        to_add[0].update ({'nb zero':nb_0,
#                                                                'beta':test['beta standard']['status'], 
#        										'beta test result': test['beta standard']['Kolmogorov test'],
#        										'beta parameters': test['beta standard']['parameters'], 
#        										'lognorm':test['lognorm']['status'], 
#        										'lognorm test result': test['lognorm']['Kolmogorov test'],
#        										'lognorm parameters': test['lognorm']['parameters'],
#                                                            'triangular':test['triangular']['status'], 
#        										'triangular test result': test['triangular']['Kolmogorov test'],
#        										'triangular parameters': test['triangular']['parameters'],
#                                                            'uniform':test['uniform']['status'], 
#        										'uniform test result': test['uniform']['Kolmogorov test'],
#        										'uniform parameters': test['uniform']['parameters']
#                                                                                                      })
#                    else :
#                        to_add[0].update ({'nb zero':1000.,'beta':'zero', 
#        										'beta test result': 'zero',
#        										'beta parameters': 'zero', 
#        										'lognorm':'zero', 
#        										'lognorm test result': 'zero',
#        										'lognorm parameters': 'zero',
#                                                            'triangular':'zero', 
#        										'triangular test result': 'zero',
#        										'triangular parameters': 'zero',
#                                                            'uniform':'zero', 
#        										'uniform test result': 'zero',
#        										'uniform parameters': 'zero'
#                                                                                                      })
#                
###                   
##                    if deterministic_CF!= 0.:
###                          to_add[0].update({'2.5/det': confidence_95[0]/deterministic_CF,
###                                            '97.5/det':confidence_95[1]/deterministic_CF,
###                                            'MC average/det':m/deterministic_CF
###									})
##                    else:
##                          to_add[0].update({'2.5/det': 'division par 0',
##                                            '97.5/det':'division par 0',
##                                            'MC average/det':'division par 0'
##									})
##                    if confidence_95[0] != 0:
##                        to_add[0].update({'97.5/2.5':confidence_95[1]/confidence_95[0]})
##                    else:
##                        to_add[0].update({'97.5/2.5':'division par 0'})
##                    if ground_surface == 'groundwater' :
##                        watertype= 'G' + water_type 
##                    else:
##                        watertype= 'S' + water_type 
##                    filename='Cell %s, %s, %s.pdf' % (cell_ID, cell_data[cell_ID]['Country'], watertype)
##                    folder = '/Users/Omar/Desktop/Projet_recherche/Water Use/Sensitivity analysis/Code'
##                elif 'NA' in CFs:
##                    to_add[0].update({'CF deterministic': 'division par 0(water availability= 0)'})  
##                elif sum(CFs)== 0.:
##                    to_add[0].update({'CF deterministic': deterministic_CF,
##                                    'alpha deterministic': deterministic_alpha, 
##									'alpha star determisnistic': deterministic_alpha_star,
##                                                       'MC average': 0., 
##									'min CF': 0.,
##                                                       'max CF': 0.,
##                                                       'min alpha stars': 0.,
##                                                       'max alpha stars' : 0.,
##                                                       'min alphas': 0.,
##                                                       'max alphas' : 0.,
##                                                       'percentile 2.5': 0.,
##                                                        'percentile 97.5':0.,
##                                                        'mediane':0.,
##                                                        '2.5/det': 'division par 0',
##                                                        '97.5/det':'division par 0',
##                                                        'MC average/det':'division par 0',
##                                                        '97.5/2.5':'division par 0'
##                                                        })                                       
#            else:
#                to_add[0].update ({'nb zero':'NA','beta':'NA', 
#        										'beta test result': 'NA',
#        										'beta parameters': 'NA', 
#        										'lognorm':'NA', 
#        										'lognorm test result': 'NA',
#        										'lognorm parameters': 'NA',
#                                                            'triangular':'NA', 
#        										'triangular test result': 'NA',
#        										'triangular parameters': 'NA',
#                                                            'uniform':'NA', 
#        										'uniform test result': 'NA',
#        										'uniform parameters': 'NA'
#                                                                                                      })
#            Kolmogorov = concat([Kolmogorov, DataFrame(to_add).transpose()]) 

# Creation des 3 excels 
        
cols = ['cell_ID', 'country', 'ground_surface', 'water_type']
cols.extend(['nb zero','beta','beta test result','beta parameters','lognorm','lognorm test result','lognorm parameters','triangular','triangular test result','triangular parameters','uniform','uniform test result','uniform parameters'])
filename = 'kolmogorov_test.xlsx'
folder = '/Users/Omar/Desktop/Projet_recherche/Water Use/Sensitivity analysis/Code'
Kolmogorov.to_excel(os.path.join(folder, filename), columns = cols, merge_cells = False)
cols = ['cell_ID', 'country', 'ground_surface', 'water_type']
cols.extend(['nombre de zero', 'anderson status','normal average', 'normal std'])
filename = 'anderson_result.xlsx'
folder = '/Users/Omar/Desktop/Projet_recherche/Water Use/Sensitivity analysis/Code'
#folder = r'C:\python\water_use_MC_2\src'
anderson_result.to_excel(os.path.join(folder, filename), columns = cols, merge_cells = False)
