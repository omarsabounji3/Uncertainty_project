#code pour l'analyse de sensibilite des seuils min et max
from openpyxl.reader.excel import load_workbook
import numpy as np
from scipy.stats import lognorm, norm, beta, uniform, anderson
from scipy.stats.mstats import spearmanr
import os
from numpy.random import dirichlet
from cPickle import load, dump
from pandas import read_excel, DataFrame, concat
from copy import deepcopy
import random as rnd 
from matplotlib.ticker import NullFormatter
import matplotlib.pyplot as mpl
import matplotlib.pyplot as plt
from numpy import linspace, multiply, log , exp
from matplotlib.pyplot import figure, show, axes,title, xlabel, ylabel
from matplotlib.backends.backend_pdf import PdfPages
import cProfile
from geo_mean import geo_mean
from sklearn import linear_model

pr = cProfile.Profile()
pr.enable()

# Fonction qui permet de lire le fichier excel

def read_data(filename):

    print 'loading ' + filename + ', takes a few seconds'

    wb = load_workbook(filename)


    ws = wb.get_sheet_by_name('Cell data')
    
    print( 'reading cell data')

    cell_data = {}

    for cell_number in range(1, 809):

        cell_data[cell_number] = {}

    row_number = 0

    header = {}

    for row in ws.rows:
    #for row in []:

        row = list(row)

        row_number += 1

        if [1, 2, 3].count(row_number) == 1:

            for column in range(len(row)):

                try:

                    row[column] = str(row[column].value)

                except TypeError:

                    row[column] = ''

            header[row_number] = row

        else:

            cell_number = float(row[0].value)

            country = str(row[1].value)

            cell_data[cell_number]['Country'] = str(row[1].value)

            for column in range(2, 7):

                tag = header[3][column]

                cell_data[cell_number][tag] = float(row[column].value)

            for column in range(7, len(row)):

                tag1 = header[1][column]

                tag2 = header[2][column]

                tag3 = header[3][column]

                try:

                    cell_data[cell_number][tag1]

                except KeyError: 

                    cell_data[cell_number][tag1] = {}

                try:

                    cell_data[cell_number][tag1][tag2]

                except KeyError: 

                    cell_data[cell_number][tag1][tag2] = {}

                try:

                    element = float(row[column].value)

                except ValueError:

                    element = str(row[column].value)
                    
                except TypeError:

                    element = str(row[column].value)
                
                cell_data[cell_number][tag1][tag2][tag3] = element

    

    ws = wb.get_sheet_by_name('Country data')
    print ('reading Country data')
    country_data = {}

    row_number = 0

    for row in ws.rows:
    #for row in []:

        row = list(row)

        row_number += 1

        if row_number != 1:

            country = str(row[0].value)

            try:

                AC = float(row[1].value)

            except ValueError:

                AC = 'NA'

            country_data[country] = {}

            country_data[country]['AC'] = AC


    

    ws = wb.get_sheet_by_name('User distribution')
    print ('reading User distribution')
    user_dist = {}

    row_number = 0

    header = {}

    for row in ws.rows:
    #for row in []:
        
        row = list(row
        )

        row_number += 1

        if [1, 2, 3].count(row_number) == 1:

            for column in range(len(row)):

                try:

                    row[column] = str(row[column].value)

                except TypeError:

                    row[column] = ''

            header[row_number] = row

        else:

            cell_number = int(row[0].value)

            user_dist[cell_number] = {}

            for column in range(1,len(row)):

                ground_surface = header[1][column]

                try:

                    user_dist[cell_number][ground_surface]

                except KeyError:

                    user_dist[cell_number][ground_surface] = {}

                water_type = header[2][column]

                try:

                    user_dist[cell_number][ground_surface][water_type]

                except KeyError:

                    user_dist[cell_number][ground_surface][water_type] = {}

                user = header[3][column]

                if user != 'Hydro' and user != 'Recre':

                    try:

                        data = float(row[column].value)

                    except TypeError:

                        data = 'NA'

                    user_dist[cell_number][ground_surface][water_type][user] = data
                    
    ws = wb.get_sheet_by_name('Uncertainty data')
    print ('reading uncertainty data')
    uncertainty_data = {}
    
    for row in ws.rows:
    #for row in []:
        
        variable = row[0]
        data_type = row[1]
        data = row[2]
    
        variable = str(variable.value)
    
        data_type = str(data_type.value)
        
        if data.value == None :  
            break
        try:
            data = float(data.value)
        except ValueError:
            
            data= str(data.value)
    
        try:
        
        
            uncertainty_data[variable]
    
        except KeyError:
        
            uncertainty_data[variable] = {}
            
        uncertainty_data[variable][data_type] = data
        
    ws = wb.get_sheet_by_name('Equation data')
    print( 'reading equation data')
    eq_data = {}
        
    for row in ws.rows:
        
        variable= row[0]
        
        data=row[1]
        
        variable = str(variable.value)
        
        data = float(data.value)
        
        eq_data[variable] = data
    
    ws = wb.get_sheet_by_name('EF')
    print ('reading EF')
    EF_data = {}
    for row in ws.rows:
    #for row in []:
        
        variable = row[0]
        data_type = row[1]
        data = row[2]
        unit = row[3]
        
        variable = str(variable.value)
        
        data_type = str(data_type.value)
        
        data = float(data.value)
        try:
            
            EF_data[variable]
        except KeyError:
            
            EF_data[variable] = {}
            
        EF_data[variable][data_type] = data

    return cell_data, country_data, user_dist, eq_data, uncertainty_data, EF_data
def save_variable(all_info):
    #all_info = [
    #           [path, variable_name, variable], 
    #           [path, variable_name, variable]
    #           ]
    for path, variable_name, variable in all_info:
        filename = os.path.join(path, variable_name + '.pkl')
        file = open(filename, 'wb')
        dump(variable, file)
        file.close()
def load_variables(all_info):
    #all_info = [[path, variable_name], [path, variable_name], ...]
    variables = []
    for path, variable_name in all_info:
        filename = path + '/' + variable_name + '.pkl'
        variables.append(load(open(filename, 'rb')))
    return variables
def build_file_list(dirpath, extension = None):
    pre_list = os.listdir(dirpath)
    if extension == None:
        filelist = [filename for filename in pre_list 
                if os.path.isfile(os.path.join(dirpath, filename)) and '~' not in filename]
    else:
        filelist = [filename for filename in pre_list 
                if os.path.isfile(os.path.join(dirpath, filename))
                and filename.split('.')[-1].lower() == extension and '~' not in filename]
    
    return filelist

# Fonction de calcul du CF 

def calculate_alpha(parameters):
    users_aggregation = {}
    users_aggregation['WR_agriculture'] = ['Agri1']
    users_aggregation['domestic'] = ['Dom1', 'Dom2', 'Dom3']
    users_aggregation['WR_fisheries'] = ['Fisheries']
    alpha = 'NA'
    alpha_star= 'NA'
    if parameters['AC'] != 1.:
        if ((parameters['SUM_GWR_K'] == 0. and ground_surface == 'groundwater') or 
                (parameters['Q_90_yr'] == 0. and ground_surface == 'surface')):
            alpha_star='NA'
        elif ground_surface == 'groundwater':
            alpha_star = parameters['CU_yr_gw']* parameters['MEAN_FG_'] / parameters['SUM_GWR_K'] / parameters['Pi']
        else:
            alpha_star = parameters['CU_yr_Surf']*(1.-parameters['MEAN_FG_'])/ parameters['Q_90_yr'] / parameters['Pi']
#            print alpha_star
        if alpha_star < parameters['min']:
            alpha = 0.
        elif alpha_star > parameters['max']:
            alpha = 1.
        else:
            #Equation du alpha star
            alpha = (alpha_star-parameters['min'])/(parameters['max'] - parameters['min'])
            #alpha = eq_data['K'] / ((1. + (eq_data['Q'] * exp(-eq_data['B'] * (alpha_star - eq_data['M'])))**(1. / eq_data['v'])))
    return alpha, alpha_star

# Creation des echantillon en fonction de la distribution du parametre

def create_rv(parameters_deterministic_values, parameters_uncertainty_values):
    RVs = {}
    all_data_available= True
    for parameter in parameters_uncertainty_values:
        if parameters_deterministic_values[parameter]==0.:
            RVs[parameter]= uniform(loc = 0., scale = 0.) 
        elif parameters_uncertainty_values[parameter]['distribution'] == 'uniform':
            if 'width' in parameters_uncertainty_values[parameter]:
                loc = parameters_deterministic_values[parameter] - parameters_uncertainty_values[parameter]['width']
                if loc< 0.:
                    1/0
                scale = 2.*parameters_uncertainty_values[parameter]['width']
            else:
                loc= parameters_uncertainty_values[parameter]['min']
                scale = parameters_uncertainty_values[parameter]['max'] -parameters_uncertainty_values[parameter]['min']
            RVs[parameter] = uniform(loc = loc, scale = scale) 
        elif parameters_uncertainty_values[parameter]['distribution'] == 'normal':
            RVs[parameter] = norm(loc = parameters_deterministic_values[parameter] , scale = parameters_uncertainty_values[parameter]['std'])
        elif parameters_uncertainty_values[parameter]['distribution'] == 'lognormal':
            RVs[parameter] = fit_lognormal(parameters_deterministic_values[parameter], parameters_uncertainty_values[parameter]['GSD2'])
        elif parameters_uncertainty_values[parameter]['distribution'] == 'beta':
#            RVs[parameter] = create_beta(parameters_deterministic_values[parameter], parameters_uncertainty_values[parameter])
            if 'width' in parameters_uncertainty_values[parameter]:
                loc = parameters_deterministic_values[parameter] - parameters_uncertainty_values[parameter]['width']
                if loc< 0.:
                    loc=0.
                scale = 2.*parameters_uncertainty_values[parameter]['width']
            else:
                loc= parameters_uncertainty_values[parameter]['min']
                scale = parameters_uncertainty_values[parameter]['max'] -parameters_uncertainty_values[parameter]['min']
            RVs[parameter] = uniform(loc = loc, scale = scale) 
    return RVs, all_data_available
# Fonction qui diagnostique la zone ou se trouve l'alpha star
def alpha_star_analysis(parameters_uncertainty_values, min_alpha_stars, max_alpha_stars):
    contents = []
    if parameters_uncertainty_values['min']['min'] > min_alpha_stars and parameters_uncertainty_values['min']['min'] < max_alpha_stars:
        contents.append(1)
    if parameters_uncertainty_values['min']['max'] > min_alpha_stars and parameters_uncertainty_values['min']['max'] < max_alpha_stars:
        contents.append(2)
    if parameters_uncertainty_values['max']['min'] > min_alpha_stars and parameters_uncertainty_values['max']['min'] < max_alpha_stars:
        contents.append(3)
    if parameters_uncertainty_values['max']['max'] > min_alpha_stars and parameters_uncertainty_values['max']['max'] < max_alpha_stars:
        contents.append(4)
    return contents
def det_alpha_star_analysis(alpha_star,eq_data):
    if alpha_star < eq_data['min_1_low']:
        diagnostic = 'alpha  = 0'
    elif alpha_star < eq_data['min_1_high'] and alpha_star > eq_data['min_1_low']:
        diagnostic = 'lower danger zone'
    elif alpha_star < eq_data['max_1_low'] and alpha_star > eq_data['min_1_high']:
        diagnostic = 'transition zone'
    elif alpha_star < eq_data['max_1_high'] and alpha_star > eq_data['max_1_low']:
        diagnostic = 'upper danger zone'
    else:
        diagnostic = 'alpha  = 1'
    return diagnostic
def fit_lognormal(average, GSD2):
    s = np.log(GSD2**.5)
    rv = lognorm(s, scale = average)
    rv.ppf([.0225, .5, .9775])
    return rv
def create_beta(average, parameters_uncertainty_values):
    if 'width' in parameters_uncertainty_values:
        minimum = average - parameters_uncertainty_values['width']
        maximum = average + parameters_uncertainty_values['width']
    else:
        minimum = parameters_uncertainty_values['min']
        maximum = parameters_uncertainty_values['max']
    if minimum < 0.:
        minimum =0.
    p975 = minimum + (maximum - minimum)*.975
    assert maximum > minimum, 'minimum should be lower than maximum'
    assert average > minimum, 'average should be higher than minimum'
    assert average < maximum, 'average should be lower than maximum'
    assert p975 > average, 'p975 should be higher than average'
    mu = (average - minimum) / (maximum - minimum)
    ba_ratio = (1 - mu)/mu
    scale = maximum - minimum
    #initial guess
    a = 1.
    b = a*ba_ratio
    threshold = .01
    direction = 0
    delta = .2
    counter = 0
    while 1:
        counter += 1
        rv = beta(a, b, loc = minimum, scale = scale)
        if abs(rv.ppf(.975)/p975) < 1.+threshold and abs(rv.ppf(.975)/p975 > 1.-threshold):
            break
        elif rv.ppf(.975) > p975:
            if direction == -1:
                delta = delta/2.
            direction = 1
        else:
            if direction == 1:
                delta = delta/2.
            direction = -1
        if counter > 1000:
            print (counter)
        a += direction*delta
        b = a*ba_ratio
    return rv
def create_sample(RVs, nb_iteration):
 #create a sample of nb_iteration for each parameter
    sample = {}
    for p in RVs:
        sample[p] = RVs[p].rvs(nb_iteration)
    return sample
def select_sample(sample, all_data, n):
    selected_data= {}
    for p in parameters_deterministic_values:
        if p in sample:
            selected_data[p]= sample[p][n]
        else:
            selected_data[p]=parameters_deterministic_values[p]
    return selected_data
def create_deltas(uncertainty_data, EF_data):
    deltas = {}
    deltas['malnutrition']= {}
    deltas['malnutrition']['min']=1./EF_data['malnutrition']['GSD2']
    deltas['malnutrition']['max']=EF_data['malnutrition']['GSD2']
    deltas['WR_fisheries']={}
    deltas['WR_fisheries']['min']=EF_data['WR_fisheries']['min']/EF_data['WR_fisheries']['average']
    deltas['WR_fisheries']['max']=EF_data['WR_fisheries']['max']/EF_data['WR_fisheries']['average']
    deltas['WR_agriculture']= {}
    deltas['WR_agriculture']['min']=EF_data['WR_agriculture']['min']/EF_data['WR_agriculture']['average']
    deltas['WR_agriculture']['max']=EF_data['WR_agriculture']['max']/EF_data['WR_agriculture']['average']
    deltas['domestic']= {}
    deltas['domestic']['min']=1./EF_data['domestic']['GSD2']
    deltas['domestic']['max']=EF_data['domestic']['GSD2']
    deltas['livestock_calories']= {}
    deltas['livestock_calories']['min']=EF_data['livestock_calories']['min']/EF_data['livestock_calories']['average']
    deltas['livestock_calories']['max']=EF_data['livestock_calories']['max']/EF_data['livestock_calories']['average']
    deltas['agri_livestock']= {}
    deltas['agri_livestock']['min']=(EF_data['agri_livestock']['average']-EF_data['agri_livestock']['width'])/EF_data['agri_livestock']['average']
    deltas['agri_livestock']['max']=(EF_data['agri_livestock']['average']+EF_data['agri_livestock']['width'])/EF_data['agri_livestock']['average']
    deltas['CU_yr_Surf']= {}
    deltas['CU_yr_Surf']['min']=1./uncertainty_data['CU_yr_Surf']['GSD2']
    deltas['CU_yr_Surf']['max']=uncertainty_data['CU_yr_Surf']['GSD2']
    deltas['CU_yr_gw']= {}
    deltas['CU_yr_gw']['min']=1./uncertainty_data['CU_yr_Surf']['GSD2']
    deltas['CU_yr_gw']['max']=uncertainty_data['CU_yr_Surf']['GSD2']
    deltas['Q_90_yr']= {}
    deltas['Q_90_yr']['min']=1./uncertainty_data['Q_90_yr']['GSD2']
    deltas['Q_90_yr']['max']=uncertainty_data['Q_90_yr']['GSD2']
    deltas['SUM_GWR_K']= {}
    deltas['SUM_GWR_K']['min']=1./uncertainty_data['SUM_GWR_K']['GSD2']
    deltas['SUM_GWR_K']['max']=uncertainty_data['SUM_GWR_K']['GSD2']
    deltas['MEAN_FG_']= {}
    deltas['MEAN_FG_']['min']=1. - uncertainty_data['MEAN_FG_']['width']
    deltas['MEAN_FG_']['max']=1. + uncertainty_data['MEAN_FG_']['width']
    deltas['Pi']= {}
    deltas['Pi']['min']=1. - uncertainty_data['Pi']['width']
    deltas['Pi']['max']=1. + uncertainty_data['Pi']['width']
    deltas['Cooling']= {}
    deltas['Cooling']['min']=1. - uncertainty_data['user_distribution']['width']
    deltas['Cooling']['max']=1. + uncertainty_data['user_distribution']['width']
    deltas['Agri1']= {}
    deltas['Agri1']['min']=1. - uncertainty_data['user_distribution']['width']
    deltas['Agri1']['max']=1. + uncertainty_data['user_distribution']['width']
    deltas['Ind']= {}
    deltas['Ind']['min']=1. - uncertainty_data['user_distribution']['width']
    deltas['Ind']['max']=1. + uncertainty_data['user_distribution']['width']
    deltas['Dom1']= {}
    deltas['Dom1']['min']=1. - uncertainty_data['user_distribution']['width']
    deltas['Dom1']['max']=1. + uncertainty_data['user_distribution']['width']
    deltas['Dom2']= {}
    deltas['Dom2']['min']=1. - uncertainty_data['user_distribution']['width']
    deltas['Dom2']['max']=1. + uncertainty_data['user_distribution']['width']
    deltas['Dom3']= {}
    deltas['Dom3']['min']=1. - uncertainty_data['user_distribution']['width']
    deltas['Dom3']['max']=1. + uncertainty_data['user_distribution']['width']
    deltas['Fisheries']= {}
    deltas['Fisheries']['min']=1. - uncertainty_data['user_distribution']['width']
    deltas['Fisheries']['max']=1. + uncertainty_data['user_distribution']['width']
#    deltas['AC']= {}
#    deltas['AC']['min']=0.9
#    deltas['AC']['max']=1.1
    deltas['max']= {}
    deltas['max']['min']=eq_data['max_1_low']/eq_data['max_1']
    deltas['max']['max']=eq_data['max_1_high']/eq_data['max_1']
    deltas['min']= {}
    deltas['min']['min']=eq_data['min_1_low']/eq_data['min_1']
    deltas['min']['max']=eq_data['min_1_high']/eq_data['min_1']
    return deltas
def extract_data(cell_ID, ground_surface, water_type, cell_data, country_data, eq_data,EF_data):
    all_data_available = False
    while 1:
        parameters = {}
        if country_data[cell_data[cell_ID]['Country']]['AC'] == 'NA':
            break
        parameters['AC'] = country_data[cell_data[cell_ID]['Country']]['AC']
        parameters['CU_yr_gw'] = cell_data[cell_ID]['CU_yr_gw']
        parameters['SUM_GWR_K'] = cell_data[cell_ID]['SUM_GWR_K']
        parameters['CU_yr_Surf'] = cell_data[cell_ID]['CU_yr_Surf']
        parameters['Q_90_yr'] = cell_data[cell_ID]['Q_90_yr']
        parameters['MEAN_FG_']= cell_data[cell_ID]['MEAN_FG_']
        if cell_data[cell_ID]['Pi'][ground_surface][water_type] in ['NA', '99999']:
            break
        parameters['Pi']=cell_data[cell_ID]['Pi'][ground_surface][water_type]
        parameters.update(user_dist[cell_ID][ground_surface][water_type])
        all_data_available = True
        break
    parameters['min'] = eq_data['min_1']
    parameters['max'] = eq_data['max_1']
    parameters['malnutrition']=EF_data['malnutrition']['average']
    parameters['domestic'] = EF_data['domestic']['average']
    parameters['WR_agriculture'] = EF_data['WR_agriculture']['average']
    parameters['WR_fisheries']= EF_data['WR_fisheries']['average']
    parameters['agri_livestock'] = EF_data['agri_livestock']['average']
    parameters['livestock_calories'] = EF_data['livestock_calories']['average']
    return parameters, all_data_available


def extract_uncertainty_info(cell_ID, ground_surface, water_type, cell_data, country_data, eq_data, EF_data, uncertainty_data):
	#put in a dictionary like this:

    parameters_uncertainty_values = {}
    parameters_uncertainty_values['malnutrition'] = {}
    parameters_uncertainty_values['domestic'] = {}
    parameters_uncertainty_values['CU_yr_Surf'] = {}
    parameters_uncertainty_values['CU_yr_gw'] = {}
    parameters_uncertainty_values['Q_90_yr'] = {}
    parameters_uncertainty_values['SUM_GWR_K'] = {}
    parameters_uncertainty_values['agri_livestock'] = {}
    parameters_uncertainty_values['MEAN_FG_'] = {}
    parameters_uncertainty_values['Pi'] = {}
    parameters_uncertainty_values['Fisheries'] = {}
    parameters_uncertainty_values['Dom1'] = {}
    parameters_uncertainty_values['Dom2'] = {}
    parameters_uncertainty_values['Dom3'] = {}
    parameters_uncertainty_values['Agri1'] = {}
#    parameters_uncertainty_values['Ind'] = {}
#    parameters_uncertainty_values['Cooling'] = {}
    parameters_uncertainty_values['min'] = {}
    parameters_uncertainty_values['max'] = {}
    parameters_uncertainty_values['WR_fisheries'] = {}
    parameters_uncertainty_values['WR_agriculture'] = {}
    parameters_uncertainty_values['livestock_calories'] = {}
    
    
    parameters_uncertainty_values['malnutrition']['GSD2'] = EF_data['malnutrition']['GSD2']
    parameters_uncertainty_values['domestic']['GSD2']=EF_data['domestic']['GSD2']
    parameters_uncertainty_values['CU_yr_Surf']['GSD2']=uncertainty_data['CU_yr_Surf']['GSD2']
    parameters_uncertainty_values['CU_yr_gw']['GSD2']=uncertainty_data['CU_yr_gw']['GSD2']
    parameters_uncertainty_values['Q_90_yr']['GSD2']=uncertainty_data['Q_90_yr']['GSD2']
    parameters_uncertainty_values['SUM_GWR_K']['GSD2']=uncertainty_data['SUM_GWR_K']['GSD2']

    parameters_uncertainty_values['agri_livestock']['width']=EF_data['agri_livestock']['width']
    parameters_uncertainty_values['MEAN_FG_']['width']=uncertainty_data['MEAN_FG_']['width']
    parameters_uncertainty_values['Pi']['width']=uncertainty_data['Pi']['width']
    parameters_uncertainty_values['Fisheries']['width']=uncertainty_data['user_distribution']['width']
    parameters_uncertainty_values['Dom1']['width']=uncertainty_data['user_distribution']['width']
    parameters_uncertainty_values['Dom2']['width']=uncertainty_data['user_distribution']['width']
    parameters_uncertainty_values['Dom3']['width']=uncertainty_data['user_distribution']['width']
    parameters_uncertainty_values['Agri1']['width']=uncertainty_data['user_distribution']['width']    
#    parameters_uncertainty_values['Ind']['width']=uncertainty_data['user_distribution']['width']
#    parameters_uncertainty_values['Cooling']['width']=uncertainty_data['user_distribution']['width']
    
    parameters_uncertainty_values['min']['min']=eq_data['min_1_low']
    parameters_uncertainty_values['min']['max']=eq_data['min_1_high']
    parameters_uncertainty_values['max']['min']=eq_data['max_1_low']
    parameters_uncertainty_values['max']['max']=eq_data['max_1_high']
    parameters_uncertainty_values['WR_fisheries']['min']=EF_data['WR_fisheries']['min']
    parameters_uncertainty_values['WR_fisheries']['max']=EF_data['WR_fisheries']['max']
    parameters_uncertainty_values['WR_agriculture']['min']=EF_data['WR_agriculture']['min']
    parameters_uncertainty_values['WR_agriculture']['max']=EF_data['WR_agriculture']['max']
    parameters_uncertainty_values['livestock_calories']['min']=EF_data['livestock_calories']['min']
    parameters_uncertainty_values['livestock_calories']['max']=EF_data['livestock_calories']['max']
    
    
    parameters_uncertainty_values['Pi']['distribution']=uncertainty_data['Pi']['distribution']
    parameters_uncertainty_values['MEAN_FG_']['distribution']=uncertainty_data['MEAN_FG_']['distribution']
    parameters_uncertainty_values['SUM_GWR_K']['distribution']=uncertainty_data['SUM_GWR_K']['distribution']
    parameters_uncertainty_values['Q_90_yr']['distribution']=uncertainty_data['Q_90_yr']['distribution']
    parameters_uncertainty_values['CU_yr_gw']['distribution']=uncertainty_data['CU_yr_gw']['distribution']
    parameters_uncertainty_values['CU_yr_Surf']['distribution']=uncertainty_data['CU_yr_Surf']['distribution']
    parameters_uncertainty_values['livestock_calories']['distribution']=uncertainty_data['livestock_calories']['distribution']
    parameters_uncertainty_values['agri_livestock']['distribution']=uncertainty_data['agri_livestock']['distribution']
    parameters_uncertainty_values['domestic']['distribution']=uncertainty_data['domestic']['distribution']
    parameters_uncertainty_values['WR_agriculture']['distribution']= uncertainty_data['WR_agriculture']['distribution']
    parameters_uncertainty_values['WR_fisheries']['distribution']=uncertainty_data['WR_fisheries']['distribution']
    parameters_uncertainty_values['malnutrition']['distribution']=uncertainty_data['malnutrition']['distribution']
    parameters_uncertainty_values['Fisheries']['distribution']=uncertainty_data['Fisheries']['distribution']
    parameters_uncertainty_values['Dom1']['distribution']=uncertainty_data['Dom1']['distribution']
    parameters_uncertainty_values['Dom2']['distribution']=uncertainty_data['Dom2']['distribution']
    parameters_uncertainty_values['Dom3']['distribution']=uncertainty_data['Dom3']['distribution']
    parameters_uncertainty_values['Agri1']['distribution']=uncertainty_data['Agri1']['distribution']    
#    parameters_uncertainty_values['Ind']['distribution']=uncertainty_data['Ind']['distribution']
#    parameters_uncertainty_values['Cooling']['distribution']=uncertainty_data['Cooling']['distribution']
    parameters_uncertainty_values['min']['distribution']=uncertainty_data['min']['distribution']
    parameters_uncertainty_values['max']['distribution']=uncertainty_data['max']['distribution']    
    

    return parameters_uncertainty_values
def variation_coeff(sample):
    coeff={}
    for p in sample:
        coeff[p]=np.std(sample[p])/np.mean(sample[p])        
    return coeff
def stats_on_alpha(alphas, alpha_stars):
    m_alpha = np.mean(alphas)
    m_alpha_star=np.mean(alpha_stars)
    confidence_95 = [np.percentile(alpha_stars, 2.5), np.percentile(alpha_stars, 97.5)]
    median = np.percentile(alpha_stars, 50.)
    min_alphas=np.min(alphas)
    min_alpha_stars=np.min(alpha_stars)
    max_alphas=np.max(alphas)
    max_alpha_stars=np.max(alpha_stars)
    return confidence_95, median, min_alpha_stars, max_alpha_stars, min_alphas, max_alphas, m_alpha,m_alpha_star 
def rel_alpha_alphastar(alpha_stars,alphas):
    fig = mpl.figure(1)
    distribution = {}
    average = {}
    spread = {}

    #for s curve
    distribution[1] = 'lognormal'
    average[1] = .03
    spread[1] = 1.7

    distribution[1] = 'uniform'
    average[1] = 10
    spread[1] = 3

    x= alpha_stars
    ar_means_x = np.mean(x)
    geo_means_x = geo_mean(x)
    domain_curve = linspace(np.min(x)*.97, np.max(x)*1.03, 1000)
    nullfmt   = NullFormatter()   

    x_proportion = .25
    y_proportion = .25
    interior_margin = .02
    exterior_margin = .1
    curve_x_width = ((1 - 2 * exterior_margin) - interior_margin) * (1. - x_proportion)
    curve_y_width = ((1 - 2*exterior_margin) - interior_margin) * (1. - y_proportion)
    histogram_x_width = ((1 - 2*exterior_margin) - interior_margin) - curve_x_width
    histogram_y_width = ((1 - 2*exterior_margin) - interior_margin) - curve_y_width

    rect_scatter = [1 - exterior_margin - curve_x_width, 
                    1 - exterior_margin - curve_y_width, 
                    curve_x_width, 
                    curve_y_width]
    rect_histx = [exterior_margin + histogram_y_width + interior_margin, 
                  exterior_margin, 
                  curve_x_width, 
                  histogram_y_width]
    rect_histy = [exterior_margin, 
                  exterior_margin + histogram_y_width + interior_margin, 
                  histogram_y_width, 
                  curve_y_width]
                  
    # start with a rectangular Figure
    fig = figure(3, figsize=(8,8))

    axScatter = axes(rect_scatter)
    axHistx = axes(rect_histx)
    axHisty = axes(rect_histy)

    # no labels
    axScatter.xaxis.set_major_formatter(nullfmt)
    axScatter.yaxis.set_major_formatter(nullfmt)

    # the scatter plot:
    operation ='s'
    if operation == 'linear':
        y = 2*x + 1
        expected_ar_mean = 2 * ar_means_x+1 
        expected_geo_mean = 2*geo_means_x+1
        curve = domain_curve* 2 + 1
        axScatter.plot(domain_curve, curve)
        axScatter.legend(('y = 2x + 1',))
    elif operation == 'quadratic':
        m = .9
        b = -.5
        y = multiply(x,x) * m + b
        expected_ar_mean = multiply(ar_means_x,ar_means_x) * m + b
        expected_geo_mean = multiply(geo_means_x, geo_means_x) * m + b
        curve = multiply(domain_curve,domain_curve) * m + b
        axScatter.plot(domain_curve, curve)
        axScatter.legend(('y = ' + str(m) + 'x^2 + ' + str(b),))
    elif operation == 'exp':
        y = exp(x)
        expected_ar_mean = exp(ar_means_x)
        expected_geo_mean = exp(geo_means_x)
        curve = exp(domain_curve)
        axScatter.plot(domain_curve, curve)
        axScatter.legend(('y = exp(x)',))
    elif operation == 's':
        #distribution[1] = 'lognormal'
        #average[1] = .03
        #spread[1] = 2.5
        parameters = {}
        parameters['min']=0.01
        parameters['max']=0.06
        low_threshold_min=0.008
        low_threshold_max=0.012
        high_threshold_min=0.04
        high_threshold_max=0.08
        y =alphas
        max_X = np.max([np.percentile(alpha_stars, 97.5),1.0])
        expected_ar_mean = (ar_means_x-parameters['min'])/(parameters['max'] - parameters['min'])
        expected_geo_mean =(geo_means_x-parameters['min'])/(parameters['max'] - parameters['min'])
        axScatter.plot([parameters['min'],parameters['max']],[0,1])
        axScatter.plot([low_threshold_min,high_threshold_max],[0,1])
        axScatter.plot([low_threshold_max,high_threshold_min,max_X],[0,1,1])

        
#                            axScatter.legend(('y = K/((1+(Q*exp(-B*(x -M)))^(1/v)))',))
    elif operation == 'ln':
        y = log(x)
        expected_ar_mean = log(ar_means_x)
        expected_geo_mean = log(geo_means_x)
        curve = log(domain_curve)
        axScatter.plot(domain_curve, curve)
        axScatter.legend(('y = ln(x)',))

    if 1:
        obtained_ar_mean = np.mean(y)
        obtained_geo_mean = geo_mean(y)
        #axScatter.set_xlim(min(x)*.97, max(x)*1.03)
        #axScatter.set_ylim(min(y)*.97, max(y)*1.03)
        #axScatter.set_xlim( (0, .1) )
        #axScatter.set_ylim( (0, 1) )
        
        n, bins, patches = axHistx.hist(x, 75, normed = 1, facecolor = 'w')
        #axHistx.plot([ar_means_x, ar_means_x], [0, max(n)*2.], 'r', linewidth = 3)
        #axHistx.plot([geo_means_x, geo_means_x], [0, max(n)*2.], 'b', linewidth = 3)
        axHistx.set_ylim((0, np.max(n)*1.05))
        axHistx.set_xlim((0, max_X))
        n, bins, patches = axHisty.hist(y, 20, orientation='horizontal', normed = 1, facecolor = 'w')
#                            n, bins, patches = axHisty.hist(y, 20, orientation='horizontal', facecolor = 'w')                           
        #axHisty.plot([0, max(n)*2.], [expected_ar_mean, expected_ar_mean], 'r', linewidth = 3)
        #axHisty.plot([0, max(n)*2.], [obtained_ar_mean, obtained_ar_mean], 'y', linewidth = 3)
        #axHisty.plot([0, max(n)*2.], [expected_geo_mean, expected_geo_mean], 'b', linewidth = 3)
        #axHisty.plot([0, max(n)*2.], [obtained_geo_mean, obtained_geo_mean], 'k', linewidth = 3)
        axHisty.set_xlim((0, 1.0))
        axHisty.set_ylim((0, 1.05))

    if operation == 'linear':
        axHistx.legend(('ar. mean: ' + str(ar_means_x)[0:6]))
        axHisty.legend(('f(ar.mean x): ', 
                        'ar.mean y: ', 
                        'f(geo.mean x): ', 
                        'geo.mean y: '), loc = 2)
        
    if 0:
        axHistx.legend(('ar. mean: ' + str(ar_means_x)[0:6], 
                        'geo. mean: ' + str(geo_means_x)[0:6]))
        axHisty.legend(('f(ar.mean x): ' + str(expected_ar_mean)[0:5], 
                        'ar.mean y: ' + str(obtained_ar_mean)[0:5], 
                        'f(geo.mean x): ' + str(expected_geo_mean)[0:5], 
                        'geo.mean y: ' + str(obtained_geo_mean)[0:5]), loc = 2)

    #axScatter.plot([ar_means_x, ar_means_x, 0], [min(curve), expected_ar_mean, expected_ar_mean], 'r', linewidth = 3)
    #axScatter.plot([geo_means_x, geo_means_x, 0], [min(curve), expected_geo_mean, expected_geo_mean], 'b', linewidth = 3)
    axScatter.set_xlim(axHistx.get_xlim())
    axScatter.set_ylim(axHisty.get_ylim())
    axHistx.set_xlabel('alpha star')
    axHisty.set_ylabel('alpha')
    #axHistx.set_xlim( axScatter.get_xlim() )
    #axHisty.set_ylim( axScatter.get_ylim() )
    for tick in axHisty.xaxis.get_major_ticks():
        tick.label1On = False
        tick.label2On = False
    for tick in axHisty.yaxis.get_major_ticks():
        tick.label1On = True
        tick.label2On = False
    for tick in axHistx.yaxis.get_major_ticks():
        tick.label1On = False
        tick.label2On = False
    show()

    return fig
def plot_histogram(data, variable_name, mean, deterministic, pdf_files, Filename,nb_iteration):
    fig = mpl.figure(1)
    ax1 = fig.add_subplot(111)
#    ax1.set_xlim(min_alphas,max_alphas*0.975)
    ax1.hist(data, 50., facecolor='grey', alpha=0.9)
    plt.axvline(x=mean, ymin=0.0, ymax = 1, linewidth=2, color='r')
    plt.axvline(x=deterministic, ymin=0.0, ymax = 1, linewidth=2, color='b')
    plt.legend(['deterministic', 'MC average'])
    ylabel('Count')
    xlabel(variable_name)
    title(Filename.replace('.pdf', ', %s iter.'% nb_iteration))
    pdf_files.savefig(fig)
    if max(data)/np.percentile(data, 97.5)> 2:
        fig = mpl.figure(1)
        ax1 = fig.add_subplot(111)
        ax1.set_xlim(min(data),np.percentile(data, 97.5))
        ax1.hist(data, 50., facecolor='purple', alpha=0.9)
        plt.axvline(x=mean, ymin=0.0, ymax = 1, linewidth=2, color='r')
        plt.axvline(x=deterministic, ymin=0.0, ymax = 1, linewidth=2, color='b')
        plt.legend(['deterministic', 'MC average'])
        ylabel('Count')
        xlabel(variable_name)
        title(Filename.replace('.pdf', ', %s iter, rescaled'% nb_iteration))
        pdf_files.savefig(fig)  
    mpl.show()
    return pdf_files
def dirichlet_sample(sample, parameters_deterministic_values):
    det_shares=[]
    CF_users_sample={}
    User={}
    User_affected={}
    User_affected_CF={}
    CF_users_positions=[]
    User['Ind']=parameters_deterministic_values['Ind']
    User['Cooling']=parameters_deterministic_values['Cooling']
    User['Dom1']=parameters_deterministic_values['Dom1']
    User['Dom2']=parameters_deterministic_values['Dom2']
    User['Dom3']=parameters_deterministic_values['Dom3']
    User['Agri1']=parameters_deterministic_values['Agri1']
    width=[]
    scaling =[1100, 100]
    for i in User:
        if User[i]!=0. :
            User_affected[i]=User[i]
    for i in User_affected:
        if i in ['Dom1','Dom2','Dom3','Agri1']:
            del sample[i]
            User_affected_CF[i]=User_affected[i]
        det_shares.append(User_affected[i]) 
    det_shares=np.array(det_shares)
    s = [dirichlet(det_shares*scaling[0], nb_iteration), 
         dirichlet(det_shares*scaling[1], nb_iteration)]
    det_shares_list=list(det_shares)
    for i in User_affected_CF:
        CF_users_positions.append(det_shares_list.index(User_affected_CF[i]))
        CF_users_sample[i]=[]
    for i in range(len(s[0])):
        for j in CF_users_positions:
            CF_users_sample[User_affected.keys()[j]].append(s[0][i][j])                
    sample.update(CF_users_sample)
    for i in CF_users_sample:
        width.append(np.percentile(CF_users_sample[i], 97.5)-np.percentile(CF_users_sample[i], 0.025))
    return sample   
pr.disable()
folder = '/Users/Omar/Desktop/Projet_recherche/Water Use/Sensitivity analysis/Code'
#folder = r'C:\python\water_use_MC_2\src'
force_read_excel=False

# Utilisation des pickles pour accelerer le calcul

if 'cell_data.pkl' in build_file_list(folder) and not force_read_excel :
    var_names = ['cell_data', 'country_data', 'user_dist','eq_data', 'uncertainty_data', 'EF_data']
    print ('loading pickle')
    all_info = zip([folder]*len(var_names), var_names)
    cell_data, country_data, user_dist,eq_data, uncertainty_data, EF_data = load_variables(all_info)
    print ('finish loading pickle' )
else:
    filename = 'water_use_all_info.xlsx'
    cell_data, country_data, user_dist,eq_data, uncertainty_data, EF_data = read_data(filename)
    all_info = [[folder, 'cell_data', cell_data], 
                [folder, 'country_data', country_data], 
                [folder, 'user_dist', user_dist], 
                [folder, 'eq_data', eq_data], 
                [folder, 'uncertainty_data', uncertainty_data], 
                [folder, 'EF_data', EF_data]]
    save_variable(all_info)
Coeff_var = DataFrame()
alpha_analysis=DataFrame()
nb_iteration = 1000
Cells={}
#for cell_ID in [1]:
#    Cells[cell_ID]={}
#    for ground_surface in ['surface']:
#        Cells[cell_ID][ground_surface]={}
#        for water_type in ['2a']:
deltas= create_deltas(uncertainty_data, EF_data)
for cell_ID in cell_data:
    Cells[cell_ID]={}
    for ground_surface in cell_data[cell_ID]['Pi']:
        Cells[cell_ID][ground_surface]={}
        for water_type in cell_data[cell_ID]['Pi'][ground_surface]:
        #calculate if all data available
            print ('Cell: %s, water: %s %s' % (cell_ID, water_type, ground_surface))
            parameters_deterministic_values, all_data_available = extract_data(cell_ID, ground_surface, water_type, cell_data, country_data, eq_data,EF_data)
            if all_data_available:
                deterministic_alpha, deterministic_alpha_star = calculate_alpha(parameters_deterministic_values)
                diagnostic=det_alpha_star_analysis(deterministic_alpha_star,eq_data)
                parameters_uncertainty_values = extract_uncertainty_info(cell_ID, ground_surface, water_type, cell_data, country_data, eq_data, EF_data, uncertainty_data)
                RVs, all_data_available = create_rv(parameters_deterministic_values, parameters_uncertainty_values)
                sample = create_sample(RVs, nb_iteration)
                sample=dirichlet_sample(sample, parameters_deterministic_values)
                alpha_stars = []
                alphas = []
                Cells[cell_ID][ground_surface][water_type]=variation_coeff(sample)
                for iteration in range(nb_iteration):        
                    selected_sample = select_sample(sample, parameters_deterministic_values, iteration)
                    alpha, alpha_star = calculate_alpha(selected_sample)
                    alpha_stars.append(alpha_star)
                    alphas.append(alpha)
                if deterministic_alpha_star!='NA':
                    confidence_95, median, min_alpha_stars, max_alpha_stars, min_alphas, max_alphas, m_alpha, m_alpha_star = stats_on_alpha(alphas, alpha_stars )                    
                    contents =alpha_star_analysis(parameters_uncertainty_values, min_alpha_stars, max_alpha_stars)
                    for p in sample:
                        to_add = {0: {'cell_ID': cell_ID, 
    										'ground_surface': ground_surface, 
    										'water_type': water_type, 
    										'country': cell_data[cell_ID]['Country'], 
    										'parameter': p, 
                                            'parameter lower rel value': deltas[p]['min'], 
                                            'parameter upper rel value': deltas[p]['max'],
    										'mean': np.mean(sample[p]),
    										'std': np.std(sample[p]), 
    										'coefficient de variation': variation_coeff(sample)[p], 
                                                                                                  }}
                        Coeff_var = concat([Coeff_var, DataFrame(to_add).transpose()]) 
                    to_add_alpha= {0: {'cell_ID': cell_ID, 
                                            'ground_surface': ground_surface, 
                                            'water_type': water_type, 
                                            'country': cell_data[cell_ID]['Country'], 
                                            'diagnostic':diagnostic,
                                            'contents': str(contents), 
                                            'alpha deterministic': deterministic_alpha, 
                                            'alpha star determisnistic': deterministic_alpha_star,
                                            'min alpha stars': min_alpha_stars,
                                            'max alpha stars' : max_alpha_stars,
                                            'min alphas': min_alphas,
                                            'max alphas' : max_alphas,
                                            'percentile 2.5': confidence_95[0],
                                            'percentile 97.5':confidence_95[1],
                                            'median':median,
                                                                                                  }}
                    alpha_analysis = concat([alpha_analysis, DataFrame(to_add_alpha).transpose()])
                    if len(alpha_stars)>0. :                                                                         
                            fig=rel_alpha_alphastar(alpha_stars,alphas)
                            filename='Cellule %s, %s, %s.pdf' % (cell_ID, cell_data[cell_ID]['Country'], water_type)
                            title(filename.replace('.pdf', ''))
                            folder = '/Users/Omar/Desktop/'
                            pdf_file = PdfPages(os.path.join(folder, filename))
                            pdf_file.savefig(fig)
                            pdf_file=plot_histogram(alpha_stars, 'alpha star',m_alpha_star, deterministic_alpha_star, pdf_file, filename,nb_iteration)
                            pdf_file=plot_histogram(alphas, 'alpha',m_alpha, deterministic_alpha, pdf_file, filename,nb_iteration)
                            pdf_file.close()                                                 
                   
# Creation des 3 excels 
        
cols = ['cell_ID', 'country', 'ground_surface', 'water_type']
cols.extend(['parameter','mean', 'std','coefficient de variation'])
filename = 'Coefficient_de_variation.xlsx'
folder = '/Users/Omar/Desktop/Projet_recherche/Water Use/Sensitivity analysis/Code'
Coeff_var.to_excel(os.path.join(folder, filename), columns = cols, merge_cells = False)

cols = ['cell_ID', 'country', 'ground_surface', 'water_type', 'diagnostic','contents']
cols.extend(['alpha deterministic', 'alpha star determisnistic'])
cols.extend(['min alpha stars','max alpha stars','min alphas','max alphas','percentile 2.5','percentile 97.5','median'])
filename = 'alpha_analysis.xlsx'
folder = '/Users/Omar/Desktop/Projet_recherche/Water Use/Sensitivity analysis/Code'
alpha_analysis.to_excel(os.path.join(folder, filename), columns = cols, merge_cells = False)